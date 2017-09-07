import pypyodbc
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment 
from openpyxl.styles import PatternFill
from openpyxl.worksheet.page import PageMargins
import datetime as dt
from excel_column_number import dict_col_converter as dcc


#Date Variable
date_cur = dt.date.today()


#Column Headers
col_header = ['Coll Code', 
              'Collector Name', 
              'Sales Rep', 
              'Sales Rep Name',  
              'Alt Company',  
              'Bus Unit', 
              'Acq Code', 
              'Link To Cust', 
              'Cust Nbr', 
              'Customer Name', 
              'Payment Terms', 
              'Pay Type', 
              'Credit Limit', 
              'Blocked', 
              'OP Period Type', 
              'AP Title', 
              'AP Name', 
              'AP Phone', 
              'AP Email', 
              'Open Balance', 
              'Not Due', 
              '0-30 Days', 
              '31-60 Days', 
              '61-90 Days', 
              '91-120 Days', 
              'Over 120 Days', 
              'Not Due', 
              'Current', 
              '0-30 Days', 
              '31-60 Days', 
              '61-90 Days', 
              '91-120 Days', 
              'Over 120 Days',
              'MTD Sales', 
              'MTD Cost', 
              'YTD Sales', 
              'YTD Cost',                
              '1st Order Date', 
              'Date', 
              'Check Nbr', 
              'Check Amount']

col_width = [19, 17, 9, 41, 12, 10, 16, 13, 9, 50, 30, 9, 11, 8, 15, 29, 45, 
             29, 64, 13, 11, 11, 11, 11, 11, 13, 11, 11, 11, 11, 11, 11, 13, 
             11, 11, 12, 12, 14, 11, 18, 14]



connection_string = '''Driver={SQL Server Native Client 11.0};
                        Server=TNDCSQL03;
                        Database=NAVREP;
                        Trusted_Connection=Yes;
                    '''

#Workbook
wb = Workbook()


#Hitouch Start---------------------------------------------------------------------------------------------------------------------------------------------------------------
sql = """
        DECLARE @AsofDate AS datetime = getdate()
        DECLARE @rdate AS datetime = @AsofDate
         
        SELECT
            (CASE WHEN CU.[Collection Agent Code] = '' THEN
                SP.[Collector Code]
            ELSE 
                CU.[Collection Agent Code]
            END) AS ColCode,
            CO.Name AS CollectorName,  
            CU.[Salesperson Code] AS SalesRepCode,
            SP.[Name] AS SalesRepName,
            CU.[Alternate Company Info Code] as AltCompy,
            CU.[Global Dimension 2 Code] as BussUnit,
            CU.[Global Dimension 1 Code] AS AcqCode,
            CU.[Link-to Customer No_] as LinkToCust,
            CU.No_ AS CustNbr,
            CU.[Name] AS CustomerName,    
            CU.[Payment Terms Code] + ' - ' + PT.Description AS PaymentTerms,
            CU.[Payment Method Code] as PayType,
            CU.[Credit Limit (LCY)] AS CredLimit,
            CASE WHEN CU.Blocked = 1 THEN
                'Yes'
            ELSE 
                'No'
            END AS Blocked,
            CU.[OP Period Type],                
            CON.[Job Title] AS APContTitle,
            CON.Name AS APContName,
            CON.[Phone No_] AS APContPhone,
            CON.[E-Mail] AS APContEmail,    
            SUM(ATB.OpenBal) AS OpenBalance,
            SUM(ATB.[Not Due]) AS NotDue,
            SUM(ATB.[0-30]) AS Days0_30,
            SUM(ATB.[31-60]) AS Days31_60,
            SUM(ATB.[61-90]) AS Days61_90,
            SUM(ATB.[91-120]) AS Days91_120,
            SUM(ATB.[Above 120]) AS Over120Days,
            SUM([D_NotDue]) AS DueNotDue,
            SUM(ATB.D_Current) AS DueCurrent,
            SUM(ATB.[D_0-30]) AS Due0_30Days,
            SUM(ATB.[D_31-60]) AS Due31_60Days,
            SUM(ATB.[D_61-90]) AS Due61_90Days,
            SUM(ATB.[D_91-120]) AS Due91_120Days,
            SUM(ATB.D_Over120) AS DueOver120Days,
            Sales.MTDSales$ AS MTDSales,
            Sales.MTDCost$ AS MTDCost,
            Sales.YTDSales$ AS YTDSales,
            Sales.YTDCost$ AS YTDCost,
            convert(date, OrdDte.First_OrdDte, 101) as '1stOrdDte',
            convert(date, DE.[Posting Date], 101) as 'LastCheckDate',    
            DE.[Document No_] AS LastCheckNbr,
            DE.Amount*-1 AS LastCheckAmount
        FROM 
        (
            SELECT
                CL.[Entry No_],
                OA.OpenBal,
                CL.[Document Date] AS AgingDate,
                0 AS [Not Due],
                CASE WHEN (CL.[Cust_ Summary Invoice No_] <> '' AND @rdate >= P.[Ending Date]) THEN
                    (CASE WHEN DATEDIFF(DAY,P.[Ending Date],@rdate) <= 30 THEN OA.OpenBal else 0 end)
                ELSE
                    (CASE WHEN DATEDIFF(DAY,CL.[Document Date],@rdate) <= 30 THEN OA.OpenBal else 0 end)
                END as [0-30],
                CASE WHEN (CL.[Cust_ Summary Invoice No_] <> '' AND @rdate >= P.[Ending Date]) THEN
                    (CASE WHEN DATEDIFF(DAY,P.[Ending Date],@rdate) between 31 and 60 THEN OA.OpenBal else 0 end)
                ELSE
                    (CASE WHEN DATEDIFF(DAY,CL.[Document Date],@rdate) between 31 and 60 THEN OA.OpenBal else 0 end)
                END as [31-60],
                CASE WHEN (CL.[Cust_ Summary Invoice No_] <> '' AND @rdate >= P.[Ending Date]) THEN
                    (CASE WHEN DATEDIFF(DAY,P.[Ending Date],@rdate) between 61 and 90 THEN OA.OpenBal else 0 end)
                ELSE
                    (CASE WHEN DATEDIFF(DAY,CL.[Document Date],@rdate) between 61 and 90 THEN OA.OpenBal else 0 end)
                END as [61-90],
                CASE WHEN (CL.[Cust_ Summary Invoice No_] <> '' AND @rdate >= P.[Ending Date]) THEN
                    (CASE WHEN DATEDIFF(DAY,P.[Ending Date],@rdate) between 91 and 120 THEN OA.OpenBal else 0 end)
                ELSE
                    (CASE WHEN DATEDIFF(DAY,CL.[Document Date],@rdate) between 91 and 120 THEN OA.OpenBal else 0 end)
                END as [91-120],
                CASE WHEN (CL.[Cust_ Summary Invoice No_] <> '' AND @rdate >= P.[Ending Date]) THEN
                    (CASE WHEN DATEDIFF(DAY,P.[Ending Date],@rdate) > 120 THEN OA.OpenBal else 0 end)
                ELSE
                    (CASE WHEN DATEDIFF(DAY,CL.[Document Date],@rdate) > 120 THEN OA.OpenBal else 0 end)
                END as [Above 120],
                0 as [D_NotDue],
                (CASE WHEN DATEDIFF(DAY,CL.[Due Date],@rdate) < 0 THEN OA.OpenBal else 0 end) as [D_Current],
                (CASE WHEN DATEDIFF(DAY,CL.[Due Date],@rdate) Between 0 and 30 THEN OA.OpenBal else 0 end) as [D_0-30],        
                (CASE WHEN DATEDIFF(DAY,CL.[Due Date],@rdate) Between 31 and 60 THEN OA.OpenBal else 0 end) as [D_31-60],
                (CASE WHEN DATEDIFF(DAY,CL.[Due Date],@rdate) Between 61 and 90 THEN OA.OpenBal else 0 end) as [D_61-90],
                (CASE WHEN DATEDIFF(DAY,CL.[Due Date],@rdate) Between 91 and 120 THEN OA.OpenBal else 0 end) as [D_91-120],
                (CASE WHEN DATEDIFF(DAY,CL.[Due Date],@rdate) > 120 THEN OA.OpenBal else 0 end) as [D_Over120]
            from
                tndcsql03.NAVRep.dbo.[Hi Touch$Cust_ Ledger Entry] CL
                left join
                    (select  
                        DL.[Cust_ Ledger Entry No_],
                        SUM(DL.Amount) as OpenBal
                    from 
                        tndcsql03.NAVRep.dbo.[Hi Touch$Detailed Cust_ Ledg_ Entry] DL
                    where 
                        DL.[Posting Date] <= @rdate
                    group by
                        DL.[Cust_ Ledger Entry No_]
                    ) OA
                        on OA.[Cust_ Ledger Entry No_] = CL.[Entry No_]
                left join tndcsql03.NAVRep.dbo.[Hi Touch$Customer Accounting Period] P
                        on CL.[Cust_ Summary Invoice No_] = P.[Cust_ Summary Invoice No_] 
                        and CL.[Customer No_] = P.[Customer No_] and CL.[Cust_ Summary Invoice No_] <> ''
            where 
                CL.[Posting Date] <= @rdate
                and 
                (CASE WHEN CL.[Cust_ Summary Invoice No_] = '' THEN 1 ELSE
                    (CASE WHEN (CL.[Cust_ Summary Invoice No_] <> '' AND P.Invoiced = 1) THEN 2 else
                        CASE WHEN CL.[Cust_ Summary Invoice No_] <> '' AND P.[Ending Date] is NULL THEN 1 
                            ELSE 3
                        END
                    END)
                end) IN (1,2)
                and ABS(OA.OpenBal) > 0
            union all
            select
                CL.[Entry No_],
                OA.OpenBal,
                P.[Ending Date] as AgingDate,
                (CASE WHEN P.Invoiced = 0 then OA.OpenBal else 0 end) as [Not Due],
                (CASE WHEN P.Invoiced = 1 and DATEDIFF(DAY,P.[Ending Date],@rdate) between 0 AND 30 THEN OA.OpenBal else 0 end) as [0-30],
                (CASE WHEN P.Invoiced = 1 and DATEDIFF(DAY,P.[Ending Date],@rdate) between 31 and 60 THEN OA.OpenBal else 0 end) as [31-60],
                (CASE WHEN P.Invoiced = 1 and DATEDIFF(DAY,P.[Ending Date],@rdate) between 61 and 90 THEN OA.OpenBal else 0 end) as [61-90],
                (CASE WHEN P.Invoiced = 1 and DATEDIFF(DAY,P.[Ending Date],@rdate) between 91 and 120 THEN OA.OpenBal else 0 end) as [91-120],
                (CASE WHEN P.Invoiced = 1 and DATEDIFF(DAY,P.[Ending Date],@rdate) > 120 THEN OA.OpenBal ELSE 0 end) as [Above 120],
                     
                --AGING by Due Date
                (CASE WHEN P.Invoiced = 0 then OA.OpenBal else 0 end) as [D_NotDue],
                0 as [D_Current],
                0 as [D_0-30],
                0 as [D_31-60],
                0 as [D_61-90],
                0 as [D_91-120],
                0  as [D_Over120]
            from
                tndcsql03.NAVRep.dbo.[Hi Touch$Cust_ Ledger Entry] CL
                left join
                (select  
                    DL.[Cust_ Ledger Entry No_],
                    SUM(DL.Amount) as OpenBal
                from 
                    tndcsql03.NAVRep.dbo.[Hi Touch$Detailed Cust_ Ledg_ Entry] DL
                where 
                    DL.[Posting Date] <= @rdate
                group by
                    DL.[Cust_ Ledger Entry No_]
                ) OA
                    on OA.[Cust_ Ledger Entry No_] = CL.[Entry No_]
                left join tndcsql03.NAVRep.dbo.[Hi Touch$Customer Accounting Period] P
                    on CL.[Cust_ Summary Invoice No_] = P.[Cust_ Summary Invoice No_] 
                    and CL.[Customer No_] = P.[Customer No_] and CL.[Cust_ Summary Invoice No_] <> ''
            where 
                CL.[Document Date] <= @rdate
                and (CASE WHEN CL.[Cust_ Summary Invoice No_] = '' THEN 1 ELSE
                    (CASE WHEN (CL.[Cust_ Summary Invoice No_] <> '' AND P.Invoiced = 1) THEN 2 else
                        CASE WHEN CL.[Cust_ Summary Invoice No_] <> '' AND P.[Ending Date] is NULL THEN 1 
                            ELSE 3
                        END
                    END)
                             
                end) IN (3)
                and ABS(OA.OpenBal) > 0
        ) ATB
            LEFT JOIN tndcsql03.NAVRep.dbo.[Hi Touch$Cust_ Ledger Entry] CL
                on CL.[Entry No_] = ATB.[Entry No_]
            LEFT JOIN tndcsql03.NAVRep.dbo.[Hi Touch$Customer] CU
                on CU.No_ = CL.[Customer No_]
            LEFT JOIN tndcsql03.NAVRep.dbo.[Hi Touch$Salesperson_Purchaser] SP with(nolock) 
                ON SP.[Code] = CU.[Salesperson Code]
            LEFT JOIN tndcsql03.NAVRep.dbo.[Hi Touch$Payment Terms] PT
                ON PT.Code = CU.[Payment Terms Code]
            LEFT JOIN tndcsql03.NAVRep.dbo.[Hi Touch$Collection Agent] CO 
                ON CO.Code = 
                        (CASE WHEN CU.[Collection Agent Code] = '' THEN SP.[Collector Code] 
                        ELSE CU.[Collection Agent Code] 
                        END)     
            LEFT JOIN tndcsql03.NAVRep.dbo.[Hi Touch$Contact] CON
                ON CON.[No_] = CU.[Primary Contact No_] 
            LEFT JOIN
                (select  
                    DL.[Cust_ Ledger Entry No_],
                    (DL.Amount) as OrigAmt
                from 
                    tndcsql03.NAVRep.dbo.[Hi Touch$Detailed Cust_ Ledg_ Entry] DL
                where 
                    DL.[Posting Date] <= @rdate
                    and DL.[Entry Type] = 1
                ) OA
                    on OA.[Cust_ Ledger Entry No_] = CL.[Entry No_]
         
            -- Last Payment
            LEFT JOIN                
                (SELECT
                    MAX(DL1.[Entry No_]) as L1ENTRY,
                    DL1.[Customer No_] as Customer_Num
                FROM     
                    tndcsql03.NAVRep.dbo.[Hi Touch$Detailed Cust_ Ledg_ Entry] DL1
                Where
                    DL1.[Document Type] = 1
                    and DL1.[Entry Type] = 1
                GROUP BY
                    DL1.[Customer No_]
                ) AS LP1
                    on LP1.Customer_Num = CL.[Customer No_]
            LEFT JOIN tndcsql03.NAVRep.dbo.[Hi Touch$Detailed Cust_ Ledg_ Entry] DE
                    ON DE.[Entry No_] = LP1.L1ENTRY
            -- MTD and YTD Sales
            LEFT JOIN
                (SELECT 
                    NavSales.CustNbr as CustNbr,    
                    Sum(CASE WHEN [Year] = Year(@rdate) and NavSales.[Month] = Month(@rdate) THEN NavSales.SaleAmount else 0 end ) as "MTDSales$", 
                    Sum(CASE WHEN [Year] = Year(@rdate) and NavSales.[Month] = Month(@rdate) THEN NavSales.Cost*1.05 else 0 end ) as "MTDCost$",
                    Sum(CASE WHEN [Year] = Year(@rdate) and NavSales.[Month] between 1 and Month(@rdate) THEN NavSales.SaleAmount else 0 end ) as "YTDSales$", 
                    Sum(CASE WHEN [Year] = Year(@rdate) and NavSales.[Month] between 1 and Month(@rdate) THEN NavSales.Cost*1.05 else 0 end ) as "YTDCost$"
                FROM     (
                        --Invoice Query
                        select 
                            H.[Sell-to Customer No_] as CustNbr,
                            Year(H.[Posting Date]) as [Year], 
                            Month(H.[Posting Date]) as [Month],
                            sum(L.Quantity*[Unit Price]) as SaleAmount,
                            sum(L.Quantity*[Unit Cost (LCY)]) as Cost
                        from 
                            tndcsql03.NAVRep.dbo.[Hi Touch$Sales Invoice Line] L
                            inner join tndcsql03.NAVRep.dbo.[Hi Touch$Sales Invoice Header] H
                                on H.[No_] = L.[Document No_]
                        group by 
                            H.[Sell-to Customer No_],Year(H.[Posting Date]), Month(H.[Posting Date])
                    union     
                        --Credit Memo Query
                        select 
                            H.[Sell-to Customer No_] as CustNbr,
                            Year(H.[Posting Date]) as [Year], 
                            Month(H.[Posting Date]) as [Month],
                            sum(L.Quantity*[Unit Price]*(-1)) as SaleAmount,
                            sum(L.Quantity*[Unit Cost (LCY)]*(-1)) as Cost
                        from 
                            tndcsql03.NAVRep.dbo.[Hi Touch$Sales Cr_Memo Line] L
                            inner join tndcsql03.NAVRep.dbo.[Hi Touch$Sales Cr_Memo Header] H
                                on H.[No_] = L.[Document No_]
                        group by 
                            H.[Sell-to Customer No_],Year(H.[Posting Date]), Month(H.[Posting Date])
                    ) NavSales
                GROUP by
                    NavSales.CustNbr
                ) Sales
                    on CL.[Customer No_] = Sales.CustNbr
            left join    --First Order Date
                (select
                    IH.[Bill-to Customer No_],
                    MIN(IH.[Order Date]) as First_OrdDte
                    from tndcsql03.NAVRep.dbo.[Hi Touch$Sales Invoice Header] IH
                where
                    IH.[Order No_] <> ' '
                    or IH.[Pre-Assigned No_] <> ' '
                group by
                    IH.[Bill-to Customer No_]
                ) OrdDte on OrdDte.[Bill-to Customer No_] = CL.[Customer No_]
        where
            CU.[Salesperson Code] not in ('1017', '1181')
        group by
            (CASE WHEN CU.[Collection Agent Code] = '' THEN
                SP.[Collector Code]
            ELSE 
                CU.[Collection Agent Code]
            END),
            CO.Name,  
            CU.[Salesperson Code],
            SP.[Name],
            CU.[Alternate Company Info Code],
            CU.[Global Dimension 2 Code],
            CU.[Global Dimension 1 Code],
            CU.[Link-to Customer No_],
            CU.No_,
            CU.[Name],    
            CU.[Payment Terms Code] + ' - ' + PT.Description,
            CU.[Payment Method Code],
            CU.[Credit Limit (LCY)],
            CASE WHEN CU.Blocked = 1 THEN
                'Yes'
            ELSE 
                'No'
            END,
            CU.[OP Period Type],                
            CON.[Job Title],
            CON.Name,
            CON.[Phone No_],
            CON.[E-Mail],    
            Sales.MTDSales$,
            Sales.MTDCost$,
            Sales.YTDSales$,
            Sales.YTDCost$,
            convert(date, OrdDte.First_OrdDte, 101),
            convert(date, DE.[Posting Date], 101),
            DE.[Document No_],
            DE.Amount*-1
        order by
            (CASE WHEN CU.[Collection Agent Code] = '' THEN
                SP.[Collector Code]
            ELSE 
                CU.[Collection Agent Code]
            END),
            CU.[Salesperson Code],
            CU.No_
    """
 
connection = pypyodbc.connect(connection_string)
cur = connection.cursor()
result = cur.execute(sql).fetchall()
cur.close()
connection.close
 
ws_cur = wb.active
ws_cur.title = 'HT'
c1 = ws_cur.cell(row=1, column=1, value='HT ATB Aged by Due Date & Document Date')
c1.font = Font(bold=True)
c1 = ws_cur.cell(row=2, column=1, value='As of {d1:%B} {d1.day}, {d1.year}'
               .format(d1=date_cur))
c1.font = Font(bold=True)
 
col_count = len(result[0])
 
row_start = 7
row_next = row_start
 
#Write column header
for x in range(0, col_count):
    c1 = ws_cur.cell(row=row_start-1, column=x+1, value=col_header[x])
    c1.font = Font(bold=True)
    c1.alignment = Alignment(horizontal='center')
     
#Format cells
#Dollars
row_next = row_start
for x in range(0, len(result)):
    c1 = ws_cur.cell(row=row_next, column=13)
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    row_next += 1
 
#Dollars    
row_next = row_start
for x in range(0, len(result)):
    for y in range(20, 38):
        c1 = ws_cur.cell(row=row_next, column=y)
        c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    row_next += 1
 
row_next = row_start
for x in range(0, len(result)):
    for y in range(41, 42):
        c1 = ws_cur.cell(row=row_next, column=y)
        c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    row_next += 1
 
#Date
row_next = row_start
for x in range(0, len(result)):
    for y in range(38, 40):
        c1 = ws_cur.cell(row=row_next, column=y)
        c1.number_format= 'm/d/yyyy'
    row_next += 1
  
#Write data
row_next = row_start
for x in range(0, len(result)):
    for y in range(0, col_count):
        c1 = ws_cur.cell(row=row_next, column=y+1, value=result[x][y])
    row_next += 1
 
 
c1 = ws_cur.cell(row=row_start, column=1)
ws_cur.freeze_panes = c1    
     
#Column Width
for y in range(0, col_count):
    ws_cur.column_dimensions[dcc.get(y+1)].width = col_width[y]
 
#Row Height        
ws_cur.row_dimensions[row_start-2].height = 2
 
#Column Labels
c1 = ws_cur.cell(row=row_start-3, column=21, value='Aged by Document Date')
ws_cur.merge_cells(start_row = row_start-3, start_column=21, 
                   end_row = row_start-3, end_column=26)
c1.alignment = Alignment(horizontal='center')
c1.font = Font(bold='true')
c1.fill = PatternFill(start_color='ffa07a', end_color='ffa07a', fill_type='solid')
     
c1 = ws_cur.cell(row=row_start-3, column=27, value='Aged by Due Date')
ws_cur.merge_cells(start_row = row_start-3, start_column=27, 
                   end_row = row_start-3, end_column=33)
c1.alignment = Alignment(horizontal='center')
c1.font = Font(bold='true')
c1.fill = PatternFill(start_color='98fb98', end_color='98fb98', fill_type='solid')    
 
c1 = ws_cur.cell(row=row_start-3, column=39, value='Last Check Payment')
ws_cur.merge_cells(start_row = row_start-3, start_column=39, 
                   end_row = row_start-3, end_column=41)
c1.alignment = Alignment(horizontal='center')
c1.font = Font(bold='true')
c1.fill = PatternFill(start_color='8470ff', end_color='8470ff', fill_type='solid')
 
#Totals
row_end = ws_cur.max_row
row_next = row_end + 2
 
for x in range(20, 38):
    c1 = ws_cur.cell(row=row_next, column=x, 
                     value='''=SUM({col_letter}{row1}:
                     {col_letter}{row2})'''.format(col_letter=dcc.get(x), 
                                                  row1=row_start, 
                                                  row2=row_end)
                     .replace('\n', ''))
    c1.font = Font(bold='true')    
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
     
     
#Page Setup    
ws_cur.page_setup.orientation = ws_cur.ORIENTATION_LANDSCAPE
ws_cur.page_setup.paper_size = ws_cur.PAPERSIZE_TABLOID
ws_cur.page_setup.fitToPage = True
ws_cur.page_setup.fitToHeight = False
ws_cur.page_setup.fitToWidth = 1
ws_cur.print_options.horizontalCentered = True
ws_cur.add_print_title(6)
ws_cur.page_margins = PageMargins(left=.5, right=.5, top=.5, bottom=.5)

#Hitouch Stop---------------------------------------------------------------------------------------------------------------------------------------------------------------














#MYOP Start---------------------------------------------------------------------------------------------------------------------------------------------------------------
sql = """
        declare @rdate date = getdate()
        DECLARE @rdateYear int = YEAR(@rdate)
        declare @rdateMonth int = Month(@rdate);
         
               
        WITH ATB AS (
                      SELECT
                             CL.[Entry No_],
                             OA.OpenBal,
                             CL.[Document Date] AS AgingDate,
                       0 AS [Not Due],
                       CASE WHEN (CL.[Cust_ Summary Invoice No_] <> '' AND @rdate >= P.[Ending Date]) THEN 
                                          (CASE WHEN DATEDIFF(DAY,P.[Ending Date],@rdate) <= 30 THEN OA.OpenBal ELSE 0 END) 
                                   ELSE
                                          (CASE WHEN DATEDIFF(DAY,CL.[Document Date],@rdate) <= 30 THEN OA.OpenBal ELSE 0 END) 
                                   END AS [0-30],
                       CASE WHEN (CL.[Cust_ Summary Invoice No_] <> '' AND @rdate >= P.[Ending Date]) THEN
                               (CASE WHEN DATEDIFF(DAY,P.[Ending Date],@rdate) between 31 AND 60 THEN OA.OpenBal ELSE 0 END) 
                                   ELSE
                               (CASE WHEN DATEDIFF(DAY,CL.[Document Date],@rdate) between 31 AND 60 THEN OA.OpenBal ELSE 0 END) 
                                 END AS [31-60],
                       CASE WHEN (CL.[Cust_ Summary Invoice No_] <> '' AND @rdate >= P.[Ending Date]) THEN
                               (CASE WHEN DATEDIFF(DAY,P.[Ending Date],@rdate) between 61 AND 90 THEN OA.OpenBal ELSE 0 END)
                           ELSE
                               (CASE WHEN DATEDIFF(DAY,CL.[Document Date],@rdate) between 61 AND 90 THEN OA.OpenBal ELSE 0 END)
                           END AS [61-90],
                       CASE WHEN (CL.[Cust_ Summary Invoice No_] <> '' AND @rdate >= P.[Ending Date]) THEN
                               (CASE WHEN DATEDIFF(DAY,P.[Ending Date],@rdate) between 91 AND 120 THEN OA.OpenBal ELSE 0 END) 
                           ELSE
                               (CASE WHEN DATEDIFF(DAY,CL.[Document Date],@rdate) between 91 AND 120 THEN OA.OpenBal ELSE 0 END) 
                                   END AS [91-120],
                       CASE WHEN (CL.[Cust_ Summary Invoice No_] <> '' AND @rdate >= P.[Ending Date]) THEN
                               (CASE WHEN DATEDIFF(DAY,P.[Ending Date],@rdate) > 120 THEN OA.OpenBal ELSE 0 END)
                           ELSE
                               (CASE WHEN DATEDIFF(DAY,CL.[Document Date],@rdate) > 120 THEN OA.OpenBal ELSE 0 END) 
                           END AS [Above 120],
                             0 AS [D_NotDue],
                             (CASE WHEN DATEDIFF(DAY,CL.[Due Date],@rdate) < 0 THEN OA.OpenBal ELSE 0 END) AS [D_Current],
                             (CASE WHEN DATEDIFF(DAY,CL.[Due Date],@rdate) Between 0 AND 30 THEN OA.OpenBal ELSE 0 END) AS [D_0-30],          
                             (CASE WHEN DATEDIFF(DAY,CL.[Due Date],@rdate) Between 31 AND 60 THEN OA.OpenBal ELSE 0 END) AS [D_31-60],
                             (CASE WHEN DATEDIFF(DAY,CL.[Due Date],@rdate) Between 61 AND 90 THEN OA.OpenBal ELSE 0 END) AS [D_61-90],
                             (CASE WHEN DATEDIFF(DAY,CL.[Due Date],@rdate) Between 91 AND 120 THEN OA.OpenBal ELSE 0 END) AS [D_91-120],
                             (CASE WHEN DATEDIFF(DAY,CL.[Due Date],@rdate) > 120 THEN OA.OpenBal ELSE 0 END) AS [D_Over120]
                      FROM
                             NAVRep.dbo.[MYOP$Cust_ Ledger Entry] CL
                             LEFT JOIN 
                                   (SELECT  
                                          DL.[Cust_ Ledger Entry No_],
                                          SUM(DL.Amount) AS OpenBal
                                   FROM 
                                          NAVRep.dbo.[MYOP$Detailed Cust_ Ledg_ Entry] DL
                                   WHERE 
                                          DL.[Posting Date] <= @rdate
                                   GROUP BY
                                          DL.[Cust_ Ledger Entry No_]
                                   ) OA ON OA.[Cust_ Ledger Entry No_] = CL.[Entry No_]
                             LEFT JOIN  NAVRep.dbo.[MYOP$Customer Accounting Period] P ON CL.[Cust_ Summary Invoice No_] = P.[Cust_ Summary Invoice No_] 
                                          AND CL.[Customer No_] = P.[Customer No_] 
                                          AND CL.[Cust_ Summary Invoice No_] <> ''
                      WHERE 
                             CL.[Posting Date] <= @rdate
                             AND (CASE WHEN CL.[Cust_ Summary Invoice No_] = '' THEN 1 
                                   ELSE (CASE WHEN (CL.[Cust_ Summary Invoice No_] <> '' AND P.Invoiced = 1) THEN 2 
                                          ELSE CASE WHEN CL.[Cust_ Summary Invoice No_] <> '' AND P.[Ending Date] IS NULL THEN 1 
                                                 ELSE 3 END
                                          END)
                                   END) IN (1,2)
                             AND ABS(OA.OpenBal) > 0
                      UNION ALL
         
                      SELECT
                             CL.[Entry No_],
                             OA.OpenBal,
                             P.[Ending Date] AS AgingDate,
                             (CASE WHEN P.Invoiced = 0 then OA.OpenBal ELSE 0 END) AS [Not Due],
                             (CASE WHEN P.Invoiced = 1 AND DATEDIFF(DAY,P.[Ending Date],@rdate) between 0 AND 30 THEN OA.OpenBal ELSE 0 END) AS [0-30],
                             (CASE WHEN P.Invoiced = 1 AND DATEDIFF(DAY,P.[Ending Date],@rdate) between 31 AND 60 THEN OA.OpenBal ELSE 0 END) AS [31-60],
                             (CASE WHEN P.Invoiced = 1 AND DATEDIFF(DAY,P.[Ending Date],@rdate) between 61 AND 90 THEN OA.OpenBal ELSE 0 END) AS [61-90],
                             (CASE WHEN P.Invoiced = 1 AND DATEDIFF(DAY,P.[Ending Date],@rdate) between 91 AND 120 THEN OA.OpenBal ELSE 0 END) AS [91-120],
                             (CASE WHEN P.Invoiced = 1 AND DATEDIFF(DAY,P.[Ending Date],@rdate) > 120 THEN OA.OpenBal ELSE 0 END) AS [Above 120],                 
                             (CASE WHEN P.Invoiced = 0 then OA.OpenBal ELSE 0 END) AS [D_NotDue],
                             0 AS [D_Current],
                             0 AS [D_0-30],
                             0 AS [D_31-60],
                             0 AS [D_61-90],
                             0 AS [D_91-120],
                             0 AS [D_Over120]
                      FROM
                             NAVRep.dbo.[MYOP$Cust_ Ledger Entry] CL
                             LEFT JOIN 
                             (SELECT  
                                   DL.[Cust_ Ledger Entry No_],
                                   SUM(DL.Amount) AS OpenBal
                             FROM 
                                   NAVRep.dbo.[MYOP$Detailed Cust_ Ledg_ Entry] DL
                             WHERE 
                                   DL.[Posting Date] <= @rdate
                             GROUP BY
                                   DL.[Cust_ Ledger Entry No_]
                             ) OA ON OA.[Cust_ Ledger Entry No_] = CL.[Entry No_]
                             LEFT JOIN  NAVRep.dbo.[MYOP$Customer Accounting Period] P ON CL.[Cust_ Summary Invoice No_] = P.[Cust_ Summary Invoice No_] 
                                   AND CL.[Customer No_] = P.[Customer No_] 
                                   AND CL.[Cust_ Summary Invoice No_] <> ''
                      WHERE 
                             CL.[Document Date] <=  @rdate
                             AND (CASE WHEN CL.[Cust_ Summary Invoice No_] = '' THEN 1 ELSE
                                          (CASE WHEN (CL.[Cust_ Summary Invoice No_] <> '' AND P.Invoiced = 1) THEN 2 ELSE
                                                 CASE WHEN CL.[Cust_ Summary Invoice No_] <> '' AND P.[Ending Date] IS NULL THEN 1 
                                                        ELSE 3 END
                                          END)
                                           
                                   END) IN (3)
                             AND ABS(OA.OpenBal) > 0
                      ) ,
               OA AS (
                
                             SELECT  
                                   DL.[Cust_ Ledger Entry No_],
                                   (DL.Amount) AS OrigAmt
                             FROM 
                                   NAVRep.dbo.[MYOP$Detailed Cust_ Ledg_ Entry] DL
                             WHERE 
                                   DL.[Posting Date] <= @rdate
                                   AND DL.[Entry Type] = 1
                         ),
               -- Last Payment
               LP1 AS (
                             SELECT
                                   MAX(DL1.[Entry No_]) AS L1ENTRY,
                                   DL1.[Customer No_] AS Customer_Num
                             FROM 
                                   NAVRep.dbo.[MYOP$Detailed Cust_ Ledg_ Entry] DL1
                             WHERE
                                   DL1.[Document Type] = 1
                                   AND DL1.[Entry Type] = 1
                             GROUP BY
                                   DL1.[Customer No_]
                             ) ,
               SALES AS (
                
                             SELECT 
                                   NavSales.CustNbr AS CustNbr,      
                                   Sum(CASE WHEN [Year] = @rdateYear AND NavSales.[Month] = @rdateMonth THEN NavSales.SaleAmount ELSE 0 END ) AS "MTDSales$", 
                                   Sum(CASE WHEN [Year] = @rdateYear AND NavSales.[Month] = @rdateMonth THEN NavSales.Cost*1.05 ELSE 0 END ) AS "MTDCost$",
                                   Sum(CASE WHEN [Year] = @rdateYear AND NavSales.[Month] between 1 AND @rdateMonth THEN NavSales.SaleAmount ELSE 0 END ) AS "YTDSales$", 
                                   Sum(CASE WHEN [Year] = @rdateYear AND NavSales.[Month] between 1 AND @rdateMonth THEN NavSales.Cost*1.05 ELSE 0 END ) AS "YTDCost$"
                             FROM   (
                                          --Invoice Query
                                          SELECT 
                                                 H.[Sell-to Customer No_] AS CustNbr,
                                                 tbd.FiscalYear AS [Year], 
                                                 tbd.FiscalMonth AS [Month],
                                                 SUM(L.Quantity*[Unit Price]) AS SaleAmount,
                                                 SUM(L.Quantity*[Unit Cost (LCY)]) AS Cost
                                          FROM 
                                                 NAVRep.dbo.[MYOP$Sales Invoice Line] L
                                                 INNER JOIN NAVRep.dbo.[MYOP$Sales Invoice Header] H ON H.[No_] = L.[Document No_]
                                                 INNER JOIN [RACDW].dbo.[t_TimeByDay] tbd ON tbd.thedate = H.[Posting Date]
                                          GROUP BY
                                                 H.[Sell-to Customer No_],tbd.FiscalYear, tbd.FiscalMonth
                                   UNION ALL     
                                          --Credit Memo Query
                                          SELECT 
                                                 H.[Sell-to Customer No_] AS CustNbr,
                                                 tbd.FiscalYear AS [Year], 
                                                 tbd.FiscalMonth AS [Month],
                                                 SUM(L.Quantity*[Unit Price]*(-1)) AS SaleAmount,
                                                 SUM(L.Quantity*[Unit Cost (LCY)]*(-1)) AS Cost
                                          FROM 
                                                 NAVRep.dbo.[MYOP$Sales Cr_Memo Line] L
                                                 INNER JOIN  NAVRep.dbo.[MYOP$Sales Cr_Memo Header] H ON H.[No_] = L.[Document No_]
                                                 INNER JOIN [RACDW].dbo.[t_TimeByDay] tbd ON tbd.[TheDate] = H.[Posting Date]
                                          GROUP BY
                                                 H.[Sell-to Customer No_],tbd.FiscalYear, tbd.FiscalMonth
                                   ) NavSales
                             GROUP by
                                   NavSales.CustNbr )
                
         
        --Query----------------------------------------------------------
                
               SELECT
                      (CASE WHEN CU.[Collection Agent Code] = '' THEN SP.[Collector Code]  ELSE CU.[Collection Agent Code] END) AS ColCode,
                      CO.Name AS CollectorName,
                      CU.[Salesperson Code] AS SalesRepCode,
                      SP.[Name] AS SalesRepName,
                      CU.[Alternate Company Info Code] AS AltCompy,   
                      CU.[Global Dimension 2 Code] AS BussUnit,       
                      CU.[Global Dimension 1 Code] AS AcqCode,
                      CU.[Link-to Customer No_] AS LinkToCust,        
                      CU.No_ AS CustNbr,
                      CU.[Name] AS CustomerName,
                      PT.Code + ' - ' + PT.Description AS PaymentTerms,
                      CU.[Payment Method Code] AS PayType,                    
                      CU.[Credit Limit (LCY)] AS CredLimit,
                      CASE WHEN CU.Blocked = 1 THEN'Yes'ELSE 'No'     END AS Blocked,
                      CU.[OP Period Type],
                      CON.[Job Title] AS APContTitle,
                      CON.Name AS APContName,
                      CON.[Phone No_] AS APContPhone,
                      CON.[E-Mail] AS APContEmail,
                      SUM(ATB.OpenBal) AS OpenBalance,
                      SUM(ATB.[Not Due]) AS NotDue,
                      SUM(ATB.[0-30]) AS Days0_30,
                      SUM(ATB.[31-60]) AS Days31_60,
                      SUM(ATB.[61-90]) AS Days61_90,
                      SUM(ATB.[91-120]) AS Days91_120,
                      SUM(ATB.[Above 120]) AS Over120Days,
                      SUM([D_NotDue]) AS DueNotDue,
                      SUM(ATB.D_Current) AS DueCurrent,
                      SUM(ATB.[D_0-30]) AS Due0_30Days,
                      SUM(ATB.[D_31-60]) AS Due31_60Days,
                      SUM(ATB.[D_61-90]) AS Due61_90Days,
                      SUM(ATB.[D_91-120]) AS Due91_120Days,
                      SUM(ATB.D_Over120) AS DueOver120Days,
                      Sales.MTDSales$ AS MTDSales,
                      Sales.MTDCost$ AS MTDCost,
                      Sales.YTDSales$ AS YTDSales,
                      Sales.YTDCost$ AS YTDCost,
                      null as '1stOrdDte',
                      convert(date, DE.[Posting Date], 101) AS LastCheckDate,
                      DE.[Document No_] AS LastCheckNbr,
                      DE.Amount*-1 AS LastCheckAmount
               FROM 
                      ATB
                      LEFT JOIN NAVRep.dbo.[MYOP$Cust_ Ledger Entry] CL ON CL.[Entry No_] = ATB.[Entry No_]
                      LEFT JOIN NAVRep.dbo.[MYOP$Customer] CU  ON CU.No_  = CL.[Customer No_]
                      LEFT JOIN NAVRep.dbo.[MYOP$Salesperson_Purchaser] SP with(nolock)    ON SP.[Code] = CU.[Salesperson Code]
                      LEFT JOIN NAVRep.dbo.[MYOP$Payment Terms] PT ON PT.Code  = CU.[Payment Terms Code]  
                      LEFT JOIN NAVRep.dbo.[MYOP$Collection Agent] CO ON CO.Code  = (CASE WHEN CU.[Collection Agent Code] = '' THEN SP.[Collector Code] ELSE CU.[Collection Agent Code] END) 
                      LEFT JOIN navrep.dbo.[MYOP$Contact] CON ON CON.[No_] Collate Latin1_General_100_CS_AS = CU.[Primary Contact No_] 
                      LEFT JOIN OA ON OA.[Cust_ Ledger Entry No_] = CL.[Entry No_]
                      LEFT JOIN LP1 ON LP1.Customer_Num = CL.[Customer No_]
                      LEFT JOIN NAVRep.dbo.[MYOP$Detailed Cust_ Ledg_ Entry] DE ON DE.[Entry No_] = LP1.L1ENTRY
                      LEFT JOIN Sales      ON CL.[Customer No_] = Sales.CustNbr     -- MTD AND YTD Sales
               where
                   CASE WHEN CU.[Collection Agent Code] = '' THEN SP.[Collector Code]  ELSE CU.[Collection Agent Code] END not in ('JENNIFER.SMITH', 'MIAONA.OSBORN')        
               GROUP BY
                      (CASE WHEN CU.[Collection Agent Code] = '' THEN SP.[Collector Code]  ELSE CU.[Collection Agent Code] END),
                      CO.Name,
                      CU.[Salesperson Code],
                      SP.[Name],
                      CU.[Alternate Company Info Code],   
                      CU.[Global Dimension 2 Code],       
                      CU.[Global Dimension 1 Code],
                      CU.[Link-to Customer No_],
                      CU.No_,
                      CU.[Name],
                      PT.Code + ' - ' + PT.Description,
                      CU.[Payment Method Code], 
                      CU.[Credit Limit (LCY)],
                      CASE WHEN CU.Blocked = 1 THEN'Yes'ELSE 'No' END,
                      CU.[OP Period Type],
                      CON.[Job Title],
                      CON.Name,
                      CON.[Phone No_],
                      CON.[E-Mail],
                      Sales.MTDSales$,
                      Sales.MTDCost$,
                      Sales.YTDSales$,
                      Sales.YTDCost$,              
                      convert(date, DE.[Posting Date], 101),
                      DE.[Document No_],
                      DE.Amount*-1
               ORDER BY
                      (CASE WHEN CU.[Collection Agent Code] = '' THEN SP.[Collector Code]  ELSE CU.[Collection Agent Code] END),
                      CU.[Salesperson Code],
              CU.No_ 
    """
 
connection = pypyodbc.connect(connection_string)
cur = connection.cursor()
result = cur.execute(sql).fetchall()
cur.close()
connection.close
 
ws_cur = wb.active
ws_cur = wb.create_sheet('MYOP')
c1 = ws_cur.cell(row=1, column=1, value='MYOP ATB Aged by Due Date & Document Date')
c1.font = Font(bold=True)
c1 = ws_cur.cell(row=2, column=1, value='As of {d1:%B} {d1.day}, {d1.year}'
               .format(d1=date_cur))
c1.font = Font(bold=True)
 
col_count = len(result[0])
 
row_start = 7
row_next = row_start
 
#Write column header
for x in range(0, col_count):
    c1 = ws_cur.cell(row=row_start-1, column=x+1, value=col_header[x])
    c1.font = Font(bold=True)
    c1.alignment = Alignment(horizontal='center')
     
#Format cells
#Dollars
row_next = row_start
for x in range(0, len(result)):
    c1 = ws_cur.cell(row=row_next, column=13)
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    row_next += 1
 
#Dollars    
row_next = row_start
for x in range(0, len(result)):
    for y in range(20, 38):
        c1 = ws_cur.cell(row=row_next, column=y)
        c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    row_next += 1
 
row_next = row_start
for x in range(0, len(result)):
    for y in range(41, 42):
        c1 = ws_cur.cell(row=row_next, column=y)
        c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    row_next += 1
 
#Date
row_next = row_start
for x in range(0, len(result)):
    for y in range(38, 40):
        c1 = ws_cur.cell(row=row_next, column=y)
        c1.number_format= 'm/d/yyyy'
    row_next += 1
  
#Write data
row_next = row_start
for x in range(0, len(result)):
    for y in range(0, col_count):
        c1 = ws_cur.cell(row=row_next, column=y+1, value=result[x][y])
    row_next += 1
 
 
c1 = ws_cur.cell(row=row_start, column=1)
ws_cur.freeze_panes = c1    
     
#Column Width
for y in range(0, col_count):
    ws_cur.column_dimensions[dcc.get(y+1)].width = col_width[y]
 
#Row Height        
ws_cur.row_dimensions[row_start-2].height = 2
 
#Column Labels
c1 = ws_cur.cell(row=row_start-3, column=21, value='Aged by Document Date')
ws_cur.merge_cells(start_row = row_start-3, start_column=21, 
                   end_row = row_start-3, end_column=26)
c1.alignment = Alignment(horizontal='center')
c1.font = Font(bold='true')
c1.fill = PatternFill(start_color='ffa07a', end_color='ffa07a', fill_type='solid')
     
c1 = ws_cur.cell(row=row_start-3, column=27, value='Aged by Due Date')
ws_cur.merge_cells(start_row = row_start-3, start_column=27, 
                   end_row = row_start-3, end_column=33)
c1.alignment = Alignment(horizontal='center')
c1.font = Font(bold='true')
c1.fill = PatternFill(start_color='98fb98', end_color='98fb98', fill_type='solid')    
 
c1 = ws_cur.cell(row=row_start-3, column=39, value='Last Check Payment')
ws_cur.merge_cells(start_row = row_start-3, start_column=39, 
                   end_row = row_start-3, end_column=41)
c1.alignment = Alignment(horizontal='center')
c1.font = Font(bold='true')
c1.fill = PatternFill(start_color='8470ff', end_color='8470ff', fill_type='solid')
 
#Totals
row_end = ws_cur.max_row
row_next = row_end + 2
 
for x in range(20, 38):
    c1 = ws_cur.cell(row=row_next, column=x, 
                     value='''=SUM({col_letter}{row1}:
                     {col_letter}{row2})'''.format(col_letter=dcc.get(x), 
                                                  row1=row_start, 
                                                  row2=row_end)
                     .replace('\n', ''))
    c1.font = Font(bold='true')    
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
     
     
#Page Setup    
ws_cur.page_setup.orientation = ws_cur.ORIENTATION_LANDSCAPE
ws_cur.page_setup.paper_size = ws_cur.PAPERSIZE_TABLOID
ws_cur.page_setup.fitToPage = True
ws_cur.page_setup.fitToHeight = False
ws_cur.page_setup.fitToWidth = 1
ws_cur.print_options.horizontalCentered = True
ws_cur.add_print_title(6)
ws_cur.page_margins = PageMargins(left=.5, right=.5, top=.5, bottom=.5)


#MYOP Stop---------------------------------------------------------------------------------------------------------------------------------------------------------------


















#RAC Start---------------------------------------------------------------------------------------------------------------------------------------------------------------
sql = """
        DECLARE @AsofDate AS datetime
        SET @AsofDate = getdate()
        
        DECLARE @rdate AS datetime
        SET @rdate = getdate()
        
        SELECT
            (CASE WHEN CU.[Collection Agent Code] = '' THEN
                SP.[Collector Code]
            ELSE CU.[Collection Agent Code]
            END) AS ColCode,
            CO.Name AS CollectorName,
            CU.[Salesperson Code] AS SalesRepCode,
            SP.[Name] AS SalesRepName,
            CU.[Alternate Company Info Code] as AltCompy,
            CU.[Global Dimension 2 Code] as BussUnit,
            CU.[Global Dimension 1 Code] AS AcqCode,
            CU.[Link-to Customer No_] as LinkToCust,
            CU.No_ AS CustNbr,
            CU.[Name] AS CustomerName,
            CU.[Payment Terms Code] + ' - ' + PT.Description AS PaymentTerms,
            CU.[Payment Method Code] as PayType,
            CU.[Credit Limit (LCY)] AS CredLimit,
            CASE WHEN CU.Blocked = 1 THEN 'Yes'
            ELSE 'No'
            END AS Blocked,
            CU.[OP Period Type],            
            CON.[Job Title] AS APContTitle,
            CON.Name As APContName,
            CON.[Phone No_] AS APContPhone,
            CON.[E-Mail] AS APContEmail,
            SUM(ATB.OpenBal) AS OpenBalance,
            SUM(ATB.[Not Due]) AS NotDue,
            SUM(ATB.[0-30]) AS Days0_30,
            SUM(ATB.[31-60]) AS Days31_60,
            SUM(ATB.[61-90]) AS Days61_90,
            SUM(ATB.[91-120]) AS Days91_120,
            SUM(ATB.[Above 120]) AS Over120Days,
            SUM([D_NotDue]) AS DueNotDue,
            SUM(ATB.D_Current) AS DueCurrent,
            SUM(ATB.[D_0-30]) AS Due0_30Days,
            SUM(ATB.[D_31-60]) AS Due31_60Days,
            SUM(ATB.[D_61-90]) AS Due61_90Days,
            SUM(ATB.[D_91-120]) AS Due91_120Days,
            SUM(ATB.D_Over120) AS DueOver120Days,
            Sales.MTDSales$ AS MTDSales,
            Sales.MTDCost$ AS MTDCost,
            Sales.YTDSales$ AS YTDSales,
            Sales.YTDCost$ AS YTDCost,
            convert(date, OrdDte.First_OrdDte, 101) as '1stOrdDte',
            convert(date, DE.[Posting Date], 101) AS LastCheckDate,
            DE.[Document No_] AS LastCheckNbr,
            DE.Amount*-1 AS LastCheckAmount
        FROM 
        (
            SELECT
                CL.[Entry No_],
                OA.OpenBal,
                CL.[Document Date] AS AgingDate,
        
                --AGING by Document Date
                0 AS [Not Due],
                CASE WHEN (CL.[Cust_ Summary Invoice No_] <> '' AND @rdate >= P.[Ending Date]) THEN
                    (CASE WHEN DATEDIFF(DAY,P.[Ending Date],@rdate) <= 30 THEN OA.OpenBal ELSE 0 END) --GP 2/13/14
        --                (CASE WHEN DATEDIFF(DAY,P.[Ending Date],@rdate -1) <= 30 THEN OA.OpenBal ELSE 0 END)--GP 2/13/14
                ELSE
                    (CASE WHEN DATEDIFF(DAY,CL.[Document Date],@rdate) <= 30 THEN OA.OpenBal ELSE 0 END) --GP 2/13/14
        --                (CASE WHEN DATEDIFF(DAY,CL.[Document Date],@rdate -1) <= 30 THEN OA.OpenBal ELSE 0 END --GP 2/13/14
                END AS [0-30],
                CASE WHEN (CL.[Cust_ Summary Invoice No_] <> '' AND @rdate >= P.[Ending Date]) THEN
                    (CASE WHEN DATEDIFF(DAY,P.[Ending Date],@rdate) BETWEEN 31 AND 60 THEN OA.OpenBal ELSE 0 END) --GP 2/13/14
        --                (CASE WHEN DATEDIFF(DAY,P.[Ending Date],@rdate -1) BETWEEN 31 AND 60 THEN OA.OpenBal ELSE 0 END) --GP 2/13/14
                ELSE
                    (CASE WHEN DATEDIFF(DAY,CL.[Document Date],@rdate) BETWEEN 31 AND 60 THEN OA.OpenBal ELSE 0 END) --GP 2/13/14 
        --                (CASE WHEN DATEDIFF(DAY,CL.[Document Date],@rdate -1) BETWEEN 31 AND 60 THEN OA.OpenBal ELSE 0 END) --GP 2/13/14 
                END AS [31-60],
                CASE WHEN (CL.[Cust_ Summary Invoice No_] <> '' AND @rdate >= P.[Ending Date]) THEN
                    (CASE WHEN DATEDIFF(DAY,P.[Ending Date],@rdate) BETWEEN 61 AND 90 THEN OA.OpenBal ELSE 0 END) --GP 2/13/14
        --              (CASE WHEN DATEDIFF(DAY,P.[Ending Date],@rdate -1) BETWEEN 61 AND 90 THEN OA.OpenBal ELSE 0 END) --GP 2/13/14
                ELSE
                    (CASE WHEN DATEDIFF(DAY,CL.[Document Date],@rdate) BETWEEN 61 AND 90 THEN OA.OpenBal ELSE 0 END) --GP 2/13/14
        --              (CASE WHEN DATEDIFF(DAY,CL.[Document Date],@rdate -1) BETWEEN 61 AND 90 THEN OA.OpenBal ELSE 0 END) --GP 2/13/14 
                END AS [61-90],
                CASE WHEN (CL.[Cust_ Summary Invoice No_] <> '' AND @rdate >= P.[Ending Date]) THEN
                    (CASE WHEN DATEDIFF(DAY,P.[Ending Date],@rdate) BETWEEN 91 AND 120 THEN OA.OpenBal ELSE 0 END) --GP 2/13/14
        --              (CASE WHEN DATEDIFF(DAY,P.[Ending Date],@rdate -1) BETWEEN 91 AND 120 THEN OA.OpenBal ELSE 0 END) --GP 2/13/14
                ELSE
                    (CASE WHEN DATEDIFF(DAY,CL.[Document Date],@rdate) BETWEEN 91 AND 120 THEN OA.OpenBal ELSE 0 END) --GP 2/13/14
        --              (CASE WHEN DATEDIFF(DAY,CL.[Document Date],@rdate -1) BETWEEN 91 AND 120 THEN OA.OpenBal ELSE 0 END)  --GP 2/13/14
                END AS [91-120],
                CASE WHEN (CL.[Cust_ Summary Invoice No_] <> '' AND @rdate >= P.[Ending Date]) THEN
                    (CASE WHEN DATEDIFF(DAY,P.[Ending Date],@rdate) > 120 THEN OA.OpenBal ELSE 0 END) --GP 2/13/14
        --              (CASE WHEN DATEDIFF(DAY,P.[Ending Date],@rdate -1) > 120 THEN OA.OpenBal ELSE 0 END) --GP 2/13/14
                ELSE
                    (CASE WHEN DATEDIFF(DAY,CL.[Document Date],@rdate) > 120 THEN OA.OpenBal ELSE 0 END) --GP 2/13/14
        --              (CASE WHEN DATEDIFF(DAY,CL.[Document Date],@rdate -1) > 120 THEN OA.OpenBal ELSE 0 END) --GP 2/13/14
                END AS [Above 120],
                0 AS [D_NotDue],
                (CASE WHEN DATEDIFF(DAY,CL.[Due Date],@rdate) < 0 THEN OA.OpenBal ELSE 0 END) AS [D_Current],
                (CASE WHEN DATEDIFF(DAY,CL.[Due Date],@rdate) BETWEEN  0 AND  30 THEN OA.OpenBal ELSE 0 END) AS [D_0-30],        
                (CASE WHEN DATEDIFF(DAY,CL.[Due Date],@rdate) BETWEEN 31 AND  60 THEN OA.OpenBal ELSE 0 END) AS [D_31-60],
                (CASE WHEN DATEDIFF(DAY,CL.[Due Date],@rdate) BETWEEN 61 AND  90 THEN OA.OpenBal ELSE 0 END) AS [D_61-90],
                (CASE WHEN DATEDIFF(DAY,CL.[Due Date],@rdate) BETWEEN 91 AND 120 THEN OA.OpenBal ELSE 0 END) AS [D_91-120],
                (CASE WHEN DATEDIFF(DAY,CL.[Due Date],@rdate) > 120 THEN OA.OpenBal ELSE 0 END) AS [D_Over120]
            FROM
                NAVRep.dbo.[Rentacrate$Cust_ Ledger Entry] CL
                LEFT JOIN
                    (SELECT  
                        DL.[Cust_ Ledger Entry No_],
                        SUM(DL.Amount) AS OpenBal
                    FROM 
                        NAVRep.dbo.[Rentacrate$Detailed Cust_ Ledg_ Entry] DL
                    WHERE 
                        DL.[Posting Date] <= @rdate
                    GROUP by
                        DL.[Cust_ Ledger Entry No_]
                    ) OA
                        ON OA.[Cust_ Ledger Entry No_] = CL.[Entry No_]
                LEFT JOIN NAVRep.dbo.[Rentacrate$Customer Accounting Period] P
                        ON CL.[Cust_ Summary Invoice No_] = P.[Cust_ Summary Invoice No_] 
                        AND CL.[Customer No_] = P.[Customer No_] AND CL.[Cust_ Summary Invoice No_] <> ''
            WHERE 
                CL.[Posting Date] <= @rdate
                AND
                (CASE WHEN CL.[Cust_ Summary Invoice No_] = '' THEN 1 ELSE
                    (CASE WHEN (CL.[Cust_ Summary Invoice No_] <> '' AND P.Invoiced = 1) THEN 2 ELSE
                        CASE WHEN CL.[Cust_ Summary Invoice No_] <> '' AND P.[Ending Date] is NULL THEN 1 ELSE 3
                        END
                    END)
                END) IN (1,2)
                AND ABS(OA.OpenBal) > 0
        
            UNION ALL
                
            SELECT
                CL.[Entry No_],
                OA.OpenBal,
                P.[Ending Date] AS AgingDate,
                (CASE WHEN P.Invoiced = 0 THEN OA.OpenBal ELSE 0 END) AS [Not Due],
                (CASE WHEN P.Invoiced = 1 AND DATEDIFF(DAY,P.[Ending Date],@rdate) BETWEEN  0 AND  30 THEN OA.OpenBal ELSE 0 END) AS [0-30],
                (CASE WHEN P.Invoiced = 1 AND DATEDIFF(DAY,P.[Ending Date],@rdate) BETWEEN 31 AND  60 THEN OA.OpenBal ELSE 0 END) AS [31-60],
                (CASE WHEN P.Invoiced = 1 AND DATEDIFF(DAY,P.[Ending Date],@rdate) BETWEEN 61 AND  90 THEN OA.OpenBal ELSE 0 END) AS [61-90],
                (CASE WHEN P.Invoiced = 1 AND DATEDIFF(DAY,P.[Ending Date],@rdate) BETWEEN 91 AND 120 THEN OA.OpenBal ELSE 0 END) AS [91-120],
                (CASE WHEN P.Invoiced = 1 AND DATEDIFF(DAY,P.[Ending Date],@rdate) > 120 THEN OA.OpenBal ELSE 0 END) AS [Above 120],
        
                --AGING by Due Date
                (CASE WHEN P.Invoiced = 0 then OA.OpenBal ELSE 0 END) AS [D_NotDue],
                0 AS [D_Current],
                0 AS [D_0-30],
                0 AS [D_31-60],
                0 AS [D_61-90],
                0 AS [D_91-120],
                0 AS [D_Over120]
            FROM
                NAVRep.dbo.[Rentacrate$Cust_ Ledger Entry] CL
                LEFT JOIN
                    (SELECT  
                        DL.[Cust_ Ledger Entry No_],
                        SUM(DL.Amount) AS OpenBal
                    FROM 
                        NAVRep.dbo.[Rentacrate$Detailed Cust_ Ledg_ Entry] DL
                    where 
                        DL.[Posting Date] <= @rdate
                    GROUP by
                        DL.[Cust_ Ledger Entry No_]
                    ) OA
                        ON OA.[Cust_ Ledger Entry No_] = CL.[Entry No_]
                LEFT JOIN NAVRep.dbo.[Rentacrate$Customer Accounting Period] P
                    ON CL.[Cust_ Summary Invoice No_] = P.[Cust_ Summary Invoice No_] 
                    AND CL.[Customer No_] = P.[Customer No_] AND CL.[Cust_ Summary Invoice No_] <> ''
            WHERE 
                CL.[Document Date] <= @rdate
                AND 
                (CASE WHEN CL.[Cust_ Summary Invoice No_] = '' THEN 1 ELSE
                    (CASE WHEN (CL.[Cust_ Summary Invoice No_] <> '' AND P.Invoiced = 1) THEN 2 ELSE
                        CASE WHEN CL.[Cust_ Summary Invoice No_] <> '' AND P.[Ending Date] is NULL THEN 1 ELSE 3
                        END
                    END)
                END IN (3))
                AND ABS(OA.OpenBal) > 0
        ) ATB
            LEFT JOIN NAVRep.dbo.[Rentacrate$Cust_ Ledger Entry] CL
                ON CL.[Entry No_] = ATB.[Entry No_]
            LEFT JOIN NAVRep.dbo.[Rentacrate$Customer] CU
                ON CU.No_ = CL.[Customer No_]
            LEFT JOIN NAVRep.dbo.[Rentacrate$Salesperson_Purchaser] SP with(nolock) 
                ON SP.[Code] = CU.[Salesperson Code]
            LEFT JOIN NAVRep.dbo.[Rentacrate$Payment Terms] PT
                ON PT.Code = CU.[Payment Terms Code]
            LEFT JOIN NAVRep.dbo.[Rentacrate$Collection Agent] CO 
                ON CO.Code = 
                        (CASE WHEN CU.[Collection Agent Code] = '' THEN SP.[Collector Code] 
                        ELSE CU.[Collection Agent Code] 
                        END) 
            LEFT JOIN NAVRep.dbo.[Rentacrate$Contact] CON
                ON CON.[No_] = CU.[Primary Contact No_]  
            LEFT JOIN
                (SELECT  
                    DL.[Cust_ Ledger Entry No_],
                    (DL.Amount) AS OrigAmt
                FROM 
                    NAVRep.dbo.[Rentacrate$Detailed Cust_ Ledg_ Entry] DL
                WHERE 
                    DL.[Posting Date] <= @rdate
                    and DL.[Entry Type] = 1
                ) OA
                    ON OA.[Cust_ Ledger Entry No_] = CL.[Entry No_]
            -- Last Payment
            LEFT JOIN                
                (SELECT
                    MAX(DL1.[Entry No_]) AS L1ENTRY,
                    DL1.[Customer No_] AS Customer_Num
                FROM     
                    NAVRep.dbo.[Rentacrate$Detailed Cust_ Ledg_ Entry] DL1
                WHERE
                    DL1.[Document Type] = 1
                    and DL1.[Entry Type] = 1
                GROUP BY
                    DL1.[Customer No_]
                ) AS LP1
                    ON LP1.Customer_Num = CL.[Customer No_]
            LEFT JOIN NAVRep.dbo.[Rentacrate$Detailed Cust_ Ledg_ Entry] DE
                    ON DE.[Entry No_] = LP1.L1ENTRY
            -- MTD and YTD Sales
            LEFT JOIN
                (SELECT 
                    NavSales.CustNbr AS CustNbr,    
                    Sum(CASE WHEN [Year] = Year(@rdate) AND NavSales.[Month] = Month(@rdate) THEN NavSales.SaleAmount ELSE 0 END ) AS "MTDSales$", 
                    Sum(CASE WHEN [Year] = Year(@rdate) AND NavSales.[Month] = Month(@rdate) THEN NavSales.Cost*1.05 ELSE 0 END )  AS "MTDCost$",
                    Sum(CASE WHEN [Year] = Year(@rdate) AND NavSales.[Month] BETWEEN 1 AND Month(@rdate) THEN NavSales.SaleAmount ELSE 0 END ) AS "YTDSales$", 
                    Sum(CASE WHEN [Year] = Year(@rdate) AND NavSales.[Month] BETWEEN 1 AND Month(@rdate) THEN NavSales.Cost*1.05 ELSE 0 END )  AS "YTDCost$"
                FROM     (
                        --Invoice Query
                        SELECT 
                            H.[Sell-to Customer No_] AS CustNbr,
                            Year(H.[Posting Date]) AS [Year], 
                            Month(H.[Posting Date]) AS [Month],
                            sum(L.Quantity*[Unit Price]) AS SaleAmount,
                            sum(L.Quantity*[Unit Cost (LCY)]) AS Cost
                        FROM 
                            NAVRep.dbo.[Rentacrate$Sales Invoice Line] L
                            INNER JOIN NAVRep.dbo.[Rentacrate$Sales Invoice Header] H
                                on H.[No_] = L.[Document No_]
                        GROUP by 
                            H.[Sell-to Customer No_],Year(H.[Posting Date]), Month(H.[Posting Date])
                    UNION     
                        --Credit Memo Query
                        SELECT 
                            H.[Sell-to Customer No_] AS CustNbr,
                            Year(H.[Posting Date]) AS [Year], 
                            Month(H.[Posting Date]) AS [Month],
                            sum(L.Quantity*[Unit Price]*(-1)) AS SaleAmount,
                            sum(L.Quantity*[Unit Cost (LCY)]*(-1)) AS Cost
                        FROM 
                            NAVRep.dbo.[Rentacrate$Sales Cr_Memo Line] L
                            INNER JOIN NAVRep.dbo.[Rentacrate$Sales Cr_Memo Header] H
                                ON H.[No_] = L.[Document No_]
                        GROUP by 
                            H.[Sell-to Customer No_],Year(H.[Posting Date]), Month(H.[Posting Date])
                    ) NavSales
                GROUP by
                    NavSales.CustNbr
                ) Sales
                    ON CL.[Customer No_] = Sales.CustNbr
            left join
                (select
                    RH.[Bill-to Customer No_],
                    MIN(RH.[Order Date]) as First_OrdDte
                    from NAVRep.dbo.[Rentacrate$Rental Header] RH
                group by
                    RH.[Bill-to Customer No_]
                ) OrdDte on OrdDte.[Bill-to Customer No_] = CL.[Customer No_]
        GROUP by
            (CASE WHEN CU.[Collection Agent Code] = '' THEN
                SP.[Collector Code]
            ELSE CU.[Collection Agent Code]
            END),
            CO.Name,
            CU.[Salesperson Code],
            SP.[Name],
            CU.[Alternate Company Info Code],
            CU.[Global Dimension 2 Code],
            CU.[Global Dimension 1 Code],
            CU.[Link-to Customer No_],
            CU.No_,
            CU.[Name],
            CU.[Payment Terms Code] + ' - ' + PT.Description,
            CU.[Payment Method Code],
            CU.[Credit Limit (LCY)],
            CASE WHEN CU.Blocked = 1 THEN 'Yes'
            ELSE 'No'
            END,
            CU.[OP Period Type],            
            CON.[Job Title],
            CON.Name,
            CON.[Phone No_],
            CON.[E-Mail],
            Sales.MTDSales$,
            Sales.MTDCost$,
            Sales.YTDSales$,
            Sales.YTDCost$,
            convert(date, OrdDte.First_OrdDte, 101),
            convert(date, DE.[Posting Date], 101),
            DE.[Document No_],
            DE.Amount*-1
        ORDER by
            (CASE WHEN CU.[Collection Agent Code] = '' THEN
                SP.[Collector Code]
            ELSE CU.[Collection Agent Code]
            END),
            CU.[Salesperson Code],
            CU.No_
    """

connection = pypyodbc.connect(connection_string)
cur = connection.cursor()
result = cur.execute(sql).fetchall()
cur.close()
connection.close

ws_cur = wb.active
ws_cur = wb.create_sheet('RAC')
c1 = ws_cur.cell(row=1, column=1, value='RAC ATB Aged by Due Date & Document Date')
c1.font = Font(bold=True)
c1 = ws_cur.cell(row=2, column=1, value='As of {d1:%B} {d1.day}, {d1.year}'
               .format(d1=date_cur))
c1.font = Font(bold=True)

col_count = len(result[0])

row_start = 7
row_next = row_start

#Write column header
for x in range(0, col_count):
    c1 = ws_cur.cell(row=row_start-1, column=x+1, value=col_header[x])
    c1.font = Font(bold=True)
    c1.alignment = Alignment(horizontal='center')
    
#Format cells
#Dollars
row_next = row_start
for x in range(0, len(result)):
    c1 = ws_cur.cell(row=row_next, column=13)
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    row_next += 1

#Dollars    
row_next = row_start
for x in range(0, len(result)):
    for y in range(20, 38):
        c1 = ws_cur.cell(row=row_next, column=y)
        c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    row_next += 1

row_next = row_start
for x in range(0, len(result)):
    for y in range(41, 42):
        c1 = ws_cur.cell(row=row_next, column=y)
        c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    row_next += 1

#Date
row_next = row_start
for x in range(0, len(result)):
    for y in range(38, 40):
        c1 = ws_cur.cell(row=row_next, column=y)
        c1.number_format= 'm/d/yyyy'
    row_next += 1
 
#Write data
row_next = row_start
for x in range(0, len(result)):
    for y in range(0, col_count):
        c1 = ws_cur.cell(row=row_next, column=y+1, value=result[x][y])
    row_next += 1


c1 = ws_cur.cell(row=row_start, column=1)
ws_cur.freeze_panes = c1    
    
#Column Width
for y in range(0, col_count):
    ws_cur.column_dimensions[dcc.get(y+1)].width = col_width[y]

#Row Height        
ws_cur.row_dimensions[row_start-2].height = 2

#Column Labels
c1 = ws_cur.cell(row=row_start-3, column=21, value='Aged by Document Date')
ws_cur.merge_cells(start_row = row_start-3, start_column=21, 
                   end_row = row_start-3, end_column=26)
c1.alignment = Alignment(horizontal='center')
c1.font = Font(bold='true')
c1.fill = PatternFill(start_color='ffa07a', end_color='ffa07a', fill_type='solid')
    
c1 = ws_cur.cell(row=row_start-3, column=27, value='Aged by Due Date')
ws_cur.merge_cells(start_row = row_start-3, start_column=27, 
                   end_row = row_start-3, end_column=33)
c1.alignment = Alignment(horizontal='center')
c1.font = Font(bold='true')
c1.fill = PatternFill(start_color='98fb98', end_color='98fb98', fill_type='solid')    

c1 = ws_cur.cell(row=row_start-3, column=39, value='Last Check Payment')
ws_cur.merge_cells(start_row = row_start-3, start_column=39, 
                   end_row = row_start-3, end_column=41)
c1.alignment = Alignment(horizontal='center')
c1.font = Font(bold='true')
c1.fill = PatternFill(start_color='8470ff', end_color='8470ff', fill_type='solid')

#Totals
row_end = ws_cur.max_row
row_next = row_end + 2

for x in range(20, 38):
    c1 = ws_cur.cell(row=row_next, column=x, 
                     value='''=SUM({col_letter}{row1}:
                     {col_letter}{row2})'''.format(col_letter=dcc.get(x), 
                                                  row1=row_start, 
                                                  row2=row_end)
                     .replace('\n', ''))
    c1.font = Font(bold='true')    
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    
    
#Page Setup    
ws_cur.page_setup.orientation = ws_cur.ORIENTATION_LANDSCAPE
ws_cur.page_setup.paper_size = ws_cur.PAPERSIZE_TABLOID
ws_cur.page_setup.fitToPage = True
ws_cur.page_setup.fitToHeight = False
ws_cur.page_setup.fitToWidth = 1
ws_cur.print_options.horizontalCentered = True
ws_cur.add_print_title(6)
ws_cur.page_margins = PageMargins(left=.5, right=.5, top=.5, bottom=.5)


#RAC Stop---------------------------------------------------------------------------------------------------------------------------------------------------------------






save_path = 'c:\\temp\\'
wb.save(save_path + 'ATB Summary {d1.year}{dmth}{dday}.xlsx'.
        format(d1=date_cur,
               dmth=str(date_cur.month).zfill(2),
               dday=str(date_cur.day).zfill(2)))
   

print('Done')

    
    

