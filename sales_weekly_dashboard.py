#Imports
import datetime as dt
import time
from dateutil.relativedelta import relativedelta as rd
from tkinter import messagebox as mb
from tkinter import Tk as tk
import sys
from time import mktime
import pypyodbc
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment 
from openpyxl.styles import PatternFill
from openpyxl.worksheet.page import PageMargins
from excel_column_number import dict_col_converter as dcc



#Date variables
yesterday = dt.date.today() + rd(days=-1)
date_start = dt.date(yesterday.year, yesterday.month, 1)
date_end = yesterday
   
#Start Date
msg_text = '''
    Use Start Date?
    {t1}    
    '''.format(t1=date_start)
 
root = tk()
root.withdraw()
answer_date = mb.askquestion(title='Date Validation', message=msg_text)
 
 
if answer_date == 'no':
    date_input = input('Type Start Date MM/DD/YYYY')
    try:
        valid_date = time.strptime(date_input, '%m/%d/%Y')
        date_start = dt.date.fromtimestamp(mktime(valid_date))                           
    except ValueError:
        print('Invalid date!')
        sys.exit()
     
#End Date
msg_text = '''
    Use End Date?
    {t1}    
    '''.format(t1=yesterday)
 
answer_date = mb.askquestion(title='Date Validation', message=msg_text)
 
 
if answer_date == 'no':
    date_input = input('Type End Date MM/DD/YYYY')
    try:
        valid_date = time.strptime(date_input, '%m/%d/%Y')
        date_end = dt.date.fromtimestamp(mktime(valid_date))                           
    except ValueError:
        print('Invalid date!')
        sys.exit()
         
#Update Sales and COGS data from DW
connection_string = '''Driver={SQL Server Native Client 11.0};
                        Server=TNDCSQL02;
                        Database=Playground;
                        Trusted_Connection=Yes;
                    '''
connection = pypyodbc.connect(connection_string)
          
cur = connection.cursor()
sql = '''
    use playground
    declare @rc int
    declare @datestart date
    declare @dateend date
    execute @rc = [myop\jason.walker].[sales_analysis_invoice_dw_monthly] 
        \'{date1}\'
        ,\'{date2}\'
    '''.format(date1=date_start, date2=date_end)
    
cur.execute(sql)
cur.commit()
cur.close()
connection.close        
    
#Dictionary Business Unit Columns
dict_bu_col = {}
dict_bu_col['DW'] = 4
dict_bu_col['FUL'] = 6
dict_bu_col['ITS'] = 8
dict_bu_col['MPS'] = 10
dict_bu_col['MYOI'] = 12
dict_bu_col['OP'] = 14

#Column Widths
w_Desc = 42
w_Amt = 12


#Hitouch Sales and Cost from DW
sql = '''
    select
        sam.Company,
        sam.BusinessUnit,
        sum(sam.SalesNoFrt) as Sales,
        sum(sam.TotalUnloadedCost) as Cost
    from Playground.[myop\jason.walker].sales_analysis_monthly_dw sam
    where
        sam.SystemName = 'NAV'
    group by
        sam.Company,
        sam.BusinessUnit
    having
        abs(sum(sam.SalesNoFrt)) > 0
        or abs(sum(sam.TotalUnloadedCost)) > 0
    '''

connection_string = '''Driver={SQL Server Native Client 11.0};
                        Server=TNDCSQL02;
                        Database=Playground;
                        Trusted_Connection=Yes;
                    '''
connection = pypyodbc.connect(connection_string)
       
cur = connection.cursor()
result = cur.execute(sql).fetchall() 
unique_company = set()

for x in result:
    unique_company.add(x[0])

cur.close()
connection.close

#Sales Analysis HT
wb = Workbook()
ws_ht = wb.active
ws_ht.title = 'HiTouch'
c = ws_ht.cell(row=1, column=1, value='Hitouch')
c = ws_ht.cell(row=2, column=1, value='Revenues and POS Gross Margins')

c = ws_ht.cell(row=3, column=1, value='For the Period {d1:%B} {d1.day}, {d1.year} to {d2:%B} {d2.day}, {d2.year}'
               .format(d1=date_start, d2=date_end))
ft1 = Font(size=12, bold=True)

for c in range(1, 4):    
    c1 = ws_ht.cell(row=c, column=1)
    c1.font = ft1
    
row_start = 7

#Company
r = row_start

for u in sorted(unique_company):
    c1 = ws_ht.cell(row=r, column=1, value =u)
    r = r + 1

for i in range(r, r+len(unique_company)):
    c1 = ws_ht.cell(row=r, column=1)
    ws_ht.column_dimensions['A'].width = w_Desc


#Total Sales, Cost
r = row_start

for u in sorted(unique_company):
    sales = sum([r[2] for r in result if r[0] == u])
    cost = sum([r[3] for r in result if r[0] == u])
    c1 = ws_ht.cell(row=r, column=2, value =sales)
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    c1 = ws_ht.cell(row=r, column=3, value =cost)
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    r = r + 1    

#Design Works Sales, Cost
r = row_start
col1 = 4
col2 = 5

for u in sorted(unique_company):
    sales = sum([r[2] for r in result if r[0] == u if r[1] == 'Design Works'])
    cost = sum([r[3] for r in result if r[0] == u if r[1] == 'Design Works'])
    c1 = ws_ht.cell(row=r, column=4, value =sales)
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    c1 = ws_ht.cell(row=r, column=5, value =cost)
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    r = r + 1

c1=ws_ht.cell(row=row_start-2, column=col1, value='Design Works')
ws_ht.merge_cells(start_row=row_start-2, 
                  start_column=col1, 
                  end_row=row_start-2, 
                  end_column=col2)
c1.alignment = Alignment(horizontal='center')
c1.font = Font(bold='true')    
c1.fill = PatternFill(start_color='eedd82', end_color='eedd82', fill_type='solid')    
    
    
#Fulfillment Sales, Cost
r = row_start
col1 = 6
col2 = 7

for u in sorted(unique_company):
    sales = sum([r[2] for r in result if r[0] == u and r[1] == 'Fulfillment'])
    cost = sum([r[3] for r in result if r[0] == u and r[1] == 'Fulfillment'])
    c1 = ws_ht.cell(row=r, column=6, value =sales)
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    c1 = ws_ht.cell(row=r, column=7, value =cost)
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    r = r + 1

c1=ws_ht.cell(row=row_start-2, column=col1, value='Fulfillment')
ws_ht.merge_cells(start_row=row_start-2, 
                  start_column=col1, 
                  end_row=row_start-2, 
                  end_column=col2)
c1.alignment = Alignment(horizontal='center')
c1.font = Font(bold='true')    
c1.fill = PatternFill(start_color='6a5acd', end_color='6a5acd', fill_type='solid')

#ITS Sales, Cost
r = row_start
col1 = 8
col2 = 9

for u in sorted(unique_company):
    sales = sum([r[2] for r in result if r[0] == u and r[1] == 'ITS'])
    cost = sum([r[3] for r in result if r[0] == u and r[1] == 'ITS'])
    c1 = ws_ht.cell(row=r, column=8, value =sales)
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    c1 = ws_ht.cell(row=r, column=9, value =cost)
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    r = r + 1

c1=ws_ht.cell(row=row_start-2, column=col1, value='ITS')
ws_ht.merge_cells(start_row=row_start-2, 
                  start_column=col1, 
                  end_row=row_start-2, 
                  end_column=col2)
c1.alignment = Alignment(horizontal='center')
c1.font = Font(bold='true')    
c1.fill = PatternFill(start_color='00bfff', end_color='00bfff', fill_type='solid')

#MPS Sales, Cost
r = row_start
col1 = 10
col2 = 11

for u in sorted(unique_company):
    sales = sum([r[2] for r in result if r[0] == u and r[1] == 'MPS'])
    cost = sum([r[3] for r in result if r[0] == u and r[1] == 'MPS'])
    c1 = ws_ht.cell(row=r, column=10, value =sales)
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    c1 = ws_ht.cell(row=r, column=11, value =cost)
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    r = r + 1

c1=ws_ht.cell(row=row_start-2, column=col1, value='MPS')
ws_ht.merge_cells(start_row=row_start-2, 
                  start_column=col1, 
                  end_row=row_start-2, 
                  end_column=col2)
c1.alignment = Alignment(horizontal='center')
c1.font = Font(bold='true')    
c1.fill = PatternFill(start_color='adff2f', end_color='adff2f', fill_type='solid')

#MYOI Sales, Cost
r = row_start
col1 = 12
col2 = 13

for u in sorted(unique_company):
    sales = sum([r[2] for r in result if r[0] == u and r[1] == 'MYOI'])
    cost = sum([r[3] for r in result if r[0] == u and r[1] == 'MYOI'])
    c1 = ws_ht.cell(row=r, column=12, value =sales)
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    c1 = ws_ht.cell(row=r, column=13, value =cost)
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    r = r + 1

c1=ws_ht.cell(row=row_start-2, column=col1, value='MYOI')
ws_ht.merge_cells(start_row=row_start-2, 
                  start_column=col1, 
                  end_row=row_start-2, 
                  end_column=col2)
c1.alignment = Alignment(horizontal='center')
c1.font = Font(bold='true')    
c1.fill = PatternFill(start_color='ff6347', end_color='ff6347', fill_type='solid')

#Office Products Sales, Cost
r = row_start
col1 = 14
col2 = 15

for u in sorted(unique_company):
    sales = sum([r[2] for r in result if r[0] == u and r[1] == 'Office Products'])
    cost = sum([r[3] for r in result if r[0] == u and r[1] == 'Office Products'])
    c1 = ws_ht.cell(row=r, column=14, value =sales)
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    c1 = ws_ht.cell(row=r, column=15, value =cost)
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    r = r + 1

c1=ws_ht.cell(row=row_start-2, column=col1, value='Office Products')
ws_ht.merge_cells(start_row=row_start-2, 
                  start_column=col1, 
                  end_row=row_start-2, 
                  end_column=col2)
c1.alignment = Alignment(horizontal='center')
c1.font = Font(bold='true')    
c1.fill = PatternFill(start_color='da70d6', end_color='da70d6', fill_type='solid')

#Sum Sales and Cost
row_end = ws_ht.max_row
col_end = ws_ht.max_column
r = row_end + 2
c1 = ws_ht.cell(row=r, column=1, value = 'DW Total')

for col in range(2, col_end+1):
    f1 = '=SUM({col_letter1}{row1}:{col_letter2}{row2})'.format(col_letter1=dcc.get(col), 
                                                                row1=row_start, 
                                                                row2=row_end, 
                                                                col_letter2=dcc.get(col))
    c1 = ws_ht.cell(row=r, column=col, value=f1)
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    ws_ht.column_dimensions['{col_letter}'.format(col_letter=
                                                  dcc.get(col))].width = w_Amt
    
#Column Headers
for c in range(2, col_end+1, 2):
    c1 = ws_ht.cell(row=row_start-1, column=c, value = 'Sales')

for c in range(3, col_end+1, 2):
    c1 = ws_ht.cell(row=row_start-1, column=c, value = 'Cost')

for c in range(2, col_end+1):
    c1 = ws_ht.cell(row=row_start-1, column=c)
    c1.alignment = Alignment(horizontal='center') 
    c1.font = Font(bold='true')


ws_ht.page_setup.orientation = ws_ht.ORIENTATION_LANDSCAPE
ws_ht.page_setup.paper_size = ws_ht.PAPERSIZE_TABLOID
ws_ht.page_setup.fitToPage = True
ws_ht.page_setup.fitToHeight = False
ws_ht.page_setup.fitToWidth = 1
ws_ht.print_options.horizontalCentered = True
ws_ht.add_print_title(6)

#Freeze Panes
c1 = ws_ht.cell(row=row_start, column=2)
ws_ht.freeze_panes = c1
    
    
   
#MYOP Sales and Cost from DW
sql = '''
    select
        sam.Company,
        sam.BusinessUnit,
        sum(sam.SalesNoFrt) as Sales,
        sum(sam.TotalUnloadedCost) as Cost
    from Playground.[myop\jason.walker].sales_analysis_monthly_dw sam
    where
        sam.SystemName = 'Vibe'
    group by
        sam.Company,
        sam.BusinessUnit
    having
        abs(sum(sam.SalesNoFrt)) > 0
        or abs(sum(sam.TotalUnloadedCost)) > 0
    '''

connection = pypyodbc.connect(connection_string)
       
cur = connection.cursor()
result = cur.execute(sql).fetchall() 
unique_company = set()

for x in result:
    unique_company.add(x[0])

cur.close()
connection.close

#Sales Analysis MYOP
ws_myop = wb.create_sheet('MYOP')
c = ws_myop.cell(row=1, column=1, value='MyOfficeProducts')
c = ws_myop.cell(row=2, column=1, value='Revenues and POS Gross Margins')

c = ws_myop.cell(row=3, column=1, value='For the Period {d1:%B} {d1.day}, {d1.year} to {d2:%B} {d2.day}, {d2.year}'
               .format(d1=date_start, d2=date_end))
ft1 = Font(size=12, bold=True)

for c in range(1, 4):    
    c1 = ws_myop.cell(row=c, column=1)
    c1.font = ft1
    
row_start = 7

#Company
r = row_start

for u in sorted(unique_company):
    c1 = ws_myop.cell(row=r, column=1, value =u)
    r = r + 1

for i in range(r, r+len(unique_company)):
    c1 = ws_myop.cell(row=r, column=1)
    ws_myop.column_dimensions['A'].width = w_Desc


#Total Sales, Cost
r = row_start

for u in sorted(unique_company):
    sales = sum([r[2] for r in result if r[0] == u])
    cost = sum([r[3] for r in result if r[0] == u])
    c1 = ws_myop.cell(row=r, column=2, value =sales)
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    c1 = ws_myop.cell(row=r, column=3, value =cost)
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    r = r + 1    

#Design Works Sales, Cost
r = row_start
col1 = 4
col2 = 5

for u in sorted(unique_company):
    sales = sum([r[2] for r in result if r[0] == u if r[1] == 'Design Works'])
    cost = sum([r[3] for r in result if r[0] == u if r[1] == 'Design Works'])
    c1 = ws_myop.cell(row=r, column=4, value =sales)
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    c1 = ws_myop.cell(row=r, column=5, value =cost)
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    r = r + 1

c1=ws_myop.cell(row=row_start-2, column=col1, value='Design Works')
ws_myop.merge_cells(start_row=row_start-2, 
                  start_column=col1, 
                  end_row=row_start-2, 
                  end_column=col2)
c1.alignment = Alignment(horizontal='center')
c1.font = Font(bold='true')    
c1.fill = PatternFill(start_color='eedd82', end_color='eedd82', fill_type='solid')    
    
    
#Fulfillment Sales, Cost
r = row_start
col1 = 6
col2 = 7

for u in sorted(unique_company):
    sales = sum([r[2] for r in result if r[0] == u and r[1] == 'Fulfillment'])
    cost = sum([r[3] for r in result if r[0] == u and r[1] == 'Fulfillment'])
    c1 = ws_myop.cell(row=r, column=6, value =sales)
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    c1 = ws_myop.cell(row=r, column=7, value =cost)
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    r = r + 1

c1=ws_myop.cell(row=row_start-2, column=col1, value='Fulfillment')
ws_myop.merge_cells(start_row=row_start-2, 
                  start_column=col1, 
                  end_row=row_start-2, 
                  end_column=col2)
c1.alignment = Alignment(horizontal='center')
c1.font = Font(bold='true')    
c1.fill = PatternFill(start_color='6a5acd', end_color='6a5acd', fill_type='solid')

#ITS Sales, Cost
r = row_start
col1 = 8
col2 = 9

for u in sorted(unique_company):
    sales = sum([r[2] for r in result if r[0] == u and r[1] == 'ITS'])
    cost = sum([r[3] for r in result if r[0] == u and r[1] == 'ITS'])
    c1 = ws_myop.cell(row=r, column=8, value =sales)
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    c1 = ws_myop.cell(row=r, column=9, value =cost)
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    r = r + 1

c1=ws_myop.cell(row=row_start-2, column=col1, value='ITS')
ws_myop.merge_cells(start_row=row_start-2, 
                  start_column=col1, 
                  end_row=row_start-2, 
                  end_column=col2)
c1.alignment = Alignment(horizontal='center')
c1.font = Font(bold='true')    
c1.fill = PatternFill(start_color='00bfff', end_color='00bfff', fill_type='solid')

#MPS Sales, Cost
r = row_start
col1 = 10
col2 = 11

for u in sorted(unique_company):
    sales = sum([r[2] for r in result if r[0] == u and r[1] == 'MPS'])
    cost = sum([r[3] for r in result if r[0] == u and r[1] == 'MPS'])
    c1 = ws_myop.cell(row=r, column=10, value =sales)
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    c1 = ws_myop.cell(row=r, column=11, value =cost)
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    r = r + 1

c1=ws_myop.cell(row=row_start-2, column=col1, value='MPS')
ws_myop.merge_cells(start_row=row_start-2, 
                  start_column=col1, 
                  end_row=row_start-2, 
                  end_column=col2)
c1.alignment = Alignment(horizontal='center')
c1.font = Font(bold='true')    
c1.fill = PatternFill(start_color='adff2f', end_color='adff2f', fill_type='solid')

#MYOI Sales, Cost
r = row_start
col1 = 12
col2 = 13

for u in sorted(unique_company):
    sales = sum([r[2] for r in result if r[0] == u and r[1] == 'MYOI'])
    cost = sum([r[3] for r in result if r[0] == u and r[1] == 'MYOI'])
    c1 = ws_myop.cell(row=r, column=12, value =sales)
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    c1 = ws_myop.cell(row=r, column=13, value =cost)
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    r = r + 1

c1=ws_myop.cell(row=row_start-2, column=col1, value='MYOI')
ws_myop.merge_cells(start_row=row_start-2, 
                  start_column=col1, 
                  end_row=row_start-2, 
                  end_column=col2)
c1.alignment = Alignment(horizontal='center')
c1.font = Font(bold='true')    
c1.fill = PatternFill(start_color='ff6347', end_color='ff6347', fill_type='solid')

#Office Products Sales, Cost
r = row_start
col1 = 14
col2 = 15

for u in sorted(unique_company):
    sales = sum([r[2] for r in result if r[0] == u and r[1] == 'Office Products'])
    cost = sum([r[3] for r in result if r[0] == u and r[1] == 'Office Products'])
    c1 = ws_myop.cell(row=r, column=14, value =sales)
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    c1 = ws_myop.cell(row=r, column=15, value =cost)
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    r = r + 1

c1=ws_myop.cell(row=row_start-2, column=col1, value='Office Products')
ws_myop.merge_cells(start_row=row_start-2, 
                  start_column=col1, 
                  end_row=row_start-2, 
                  end_column=col2)
c1.alignment = Alignment(horizontal='center')
c1.font = Font(bold='true')    
c1.fill = PatternFill(start_color='da70d6', end_color='da70d6', fill_type='solid')

#Sum Sales and Cost
row_end = ws_myop.max_row
col_end = ws_myop.max_column
r = row_end + 2
c1 = ws_myop.cell(row=r, column=1, value = 'DW Total')

for col in range(2, col_end+1):
    f1 = '=SUM({col_letter1}{row1}:{col_letter2}{row2})'.format(col_letter1=dcc.get(col), 
                                                                row1=row_start, 
                                                                row2=row_end, 
                                                                col_letter2=dcc.get(col))
    c1 = ws_myop.cell(row=r, column=col, value=f1)
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    ws_myop.column_dimensions['{col_letter}'.format(col_letter=
                                                  dcc.get(col))].width = w_Amt
    
#Column Headers
for c in range(2, col_end+1, 2):
    c1 = ws_myop.cell(row=row_start-1, column=c, value = 'Sales')

for c in range(3, col_end+1, 2):
    c1 = ws_myop.cell(row=row_start-1, column=c, value = 'Cost')

for c in range(2, col_end+1):
    c1 = ws_myop.cell(row=row_start-1, column=c)
    c1.alignment = Alignment(horizontal='center') 
    c1.font = Font(bold='true')


ws_myop.page_setup.orientation = ws_myop.ORIENTATION_LANDSCAPE
ws_myop.page_setup.paper_size = ws_myop.PAPERSIZE_TABLOID
ws_myop.page_setup.fitToPage = True
ws_myop.page_setup.fitToHeight = False
ws_myop.page_setup.fitToWidth = 1
ws_myop.print_options.horizontalCentered = True
ws_myop.add_print_title(6)

#Freeze Panes
c1 = ws_myop.cell(row=row_start, column=2)
ws_myop.freeze_panes = c1    




#HiTouch MYOP Sales and Cost from DW
sql = '''
    select
        sam.Company,
        sam.BusinessUnit,
        sum(sam.SalesNoFrt) as Sales,
        sum(sam.TotalUnloadedCost) as Cost
    from Playground.[myop\jason.walker].sales_analysis_monthly_dw sam
    group by
        sam.Company,
        sam.BusinessUnit
    having
        abs(sum(sam.SalesNoFrt)) > 0
        or abs(sum(sam.TotalUnloadedCost)) > 0
    '''

connection = pypyodbc.connect(connection_string)
       
cur = connection.cursor()
result = cur.execute(sql).fetchall() 
unique_company = set()

for x in result:
    unique_company.add(x[0])

cur.close()
connection.close

#Sales Analysis HiTouch and MYOP
ws_both = wb.create_sheet('HiTouch_and_MYOP')
c = ws_both.cell(row=1, column=1, value='Hitouch Business Services')
c = ws_both.cell(row=2, column=1, value='Revenues and POS Gross Margins')

c = ws_both.cell(row=3, column=1, value='For the Period {d1:%B} {d1.day}, {d1.year} to {d2:%B} {d2.day}, {d2.year}'
               .format(d1=date_start, d2=date_end))
ft1 = Font(size=12, bold=True)

for c in range(1, 4):    
    c1 = ws_both.cell(row=c, column=1)
    c1.font = ft1
    
row_start = 7

#Company
r = row_start

for u in sorted(unique_company):
    c1 = ws_both.cell(row=r, column=1, value =u)
    r = r + 1

for i in range(r, r+len(unique_company)):
    c1 = ws_both.cell(row=r, column=1)
    ws_both.column_dimensions['A'].width = w_Desc


#Total Sales, Cost
r = row_start

for u in sorted(unique_company):
    sales = sum([r[2] for r in result if r[0] == u])
    cost = sum([r[3] for r in result if r[0] == u])
    c1 = ws_both.cell(row=r, column=2, value =sales)
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    c1 = ws_both.cell(row=r, column=3, value =cost)
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    r = r + 1    

#Design Works Sales, Cost
r = row_start
col1 = 4
col2 = 5

for u in sorted(unique_company):
    sales = sum([r[2] for r in result if r[0] == u if r[1] == 'Design Works'])
    cost = sum([r[3] for r in result if r[0] == u if r[1] == 'Design Works'])
    c1 = ws_both.cell(row=r, column=4, value =sales)
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    c1 = ws_both.cell(row=r, column=5, value =cost)
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    r = r + 1

c1=ws_both.cell(row=row_start-2, column=col1, value='Design Works')
ws_both.merge_cells(start_row=row_start-2, 
                  start_column=col1, 
                  end_row=row_start-2, 
                  end_column=col2)
c1.alignment = Alignment(horizontal='center')
c1.font = Font(bold='true')    
c1.fill = PatternFill(start_color='eedd82', end_color='eedd82', fill_type='solid')    
    
    
#Fulfillment Sales, Cost
r = row_start
col1 = 6
col2 = 7

for u in sorted(unique_company):
    sales = sum([r[2] for r in result if r[0] == u and r[1] == 'Fulfillment'])
    cost = sum([r[3] for r in result if r[0] == u and r[1] == 'Fulfillment'])
    c1 = ws_both.cell(row=r, column=6, value =sales)
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    c1 = ws_both.cell(row=r, column=7, value =cost)
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    r = r + 1

c1=ws_both.cell(row=row_start-2, column=col1, value='Fulfillment')
ws_both.merge_cells(start_row=row_start-2, 
                  start_column=col1, 
                  end_row=row_start-2, 
                  end_column=col2)
c1.alignment = Alignment(horizontal='center')
c1.font = Font(bold='true')    
c1.fill = PatternFill(start_color='6a5acd', end_color='6a5acd', fill_type='solid')

#ITS Sales, Cost
r = row_start
col1 = 8
col2 = 9

for u in sorted(unique_company):
    sales = sum([r[2] for r in result if r[0] == u and r[1] == 'ITS'])
    cost = sum([r[3] for r in result if r[0] == u and r[1] == 'ITS'])
    c1 = ws_both.cell(row=r, column=8, value =sales)
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    c1 = ws_both.cell(row=r, column=9, value =cost)
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    r = r + 1

c1=ws_both.cell(row=row_start-2, column=col1, value='ITS')
ws_both.merge_cells(start_row=row_start-2, 
                  start_column=col1, 
                  end_row=row_start-2, 
                  end_column=col2)
c1.alignment = Alignment(horizontal='center')
c1.font = Font(bold='true')    
c1.fill = PatternFill(start_color='00bfff', end_color='00bfff', fill_type='solid')

#MPS Sales, Cost
r = row_start
col1 = 10
col2 = 11

for u in sorted(unique_company):
    sales = sum([r[2] for r in result if r[0] == u and r[1] == 'MPS'])
    cost = sum([r[3] for r in result if r[0] == u and r[1] == 'MPS'])
    c1 = ws_both.cell(row=r, column=10, value =sales)
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    c1 = ws_both.cell(row=r, column=11, value =cost)
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    r = r + 1

c1=ws_both.cell(row=row_start-2, column=col1, value='MPS')
ws_both.merge_cells(start_row=row_start-2, 
                  start_column=col1, 
                  end_row=row_start-2, 
                  end_column=col2)
c1.alignment = Alignment(horizontal='center')
c1.font = Font(bold='true')    
c1.fill = PatternFill(start_color='adff2f', end_color='adff2f', fill_type='solid')

#MYOI Sales, Cost
r = row_start
col1 = 12
col2 = 13

for u in sorted(unique_company):
    sales = sum([r[2] for r in result if r[0] == u and r[1] == 'MYOI'])
    cost = sum([r[3] for r in result if r[0] == u and r[1] == 'MYOI'])
    c1 = ws_both.cell(row=r, column=12, value =sales)
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    c1 = ws_both.cell(row=r, column=13, value =cost)
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    r = r + 1

c1=ws_both.cell(row=row_start-2, column=col1, value='MYOI')
ws_both.merge_cells(start_row=row_start-2, 
                  start_column=col1, 
                  end_row=row_start-2, 
                  end_column=col2)
c1.alignment = Alignment(horizontal='center')
c1.font = Font(bold='true')    
c1.fill = PatternFill(start_color='ff6347', end_color='ff6347', fill_type='solid')

#Office Products Sales, Cost
r = row_start
col1 = 14
col2 = 15

for u in sorted(unique_company):
    sales = sum([r[2] for r in result if r[0] == u and r[1] == 'Office Products'])
    cost = sum([r[3] for r in result if r[0] == u and r[1] == 'Office Products'])
    c1 = ws_both.cell(row=r, column=14, value =sales)
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    c1 = ws_both.cell(row=r, column=15, value =cost)
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    r = r + 1

c1=ws_both.cell(row=row_start-2, column=col1, value='Office Products')
ws_both.merge_cells(start_row=row_start-2, 
                  start_column=col1, 
                  end_row=row_start-2, 
                  end_column=col2)
c1.alignment = Alignment(horizontal='center')
c1.font = Font(bold='true')    
c1.fill = PatternFill(start_color='da70d6', end_color='da70d6', fill_type='solid')

#Sum Sales and Cost
row_end = ws_both.max_row
col_end = ws_both.max_column
r = row_end + 2
c1 = ws_both.cell(row=r, column=1, value = 'DW Total')

for col in range(2, col_end+1):
    f1 = '=SUM({col_letter1}{row1}:{col_letter2}{row2})'.format(col_letter1=dcc.get(col), 
                                                                row1=row_start, 
                                                                row2=row_end, 
                                                                col_letter2=dcc.get(col))
    c1 = ws_both.cell(row=r, column=col, value=f1)
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    ws_both.column_dimensions['{col_letter}'.format(col_letter=
                                                  dcc.get(col))].width = w_Amt
    
#Column Headers
for c in range(2, col_end+1, 2):
    c1 = ws_both.cell(row=row_start-1, column=c, value = 'Sales')

for c in range(3, col_end+1, 2):
    c1 = ws_both.cell(row=row_start-1, column=c, value = 'Cost')

for c in range(2, col_end+1):
    c1 = ws_both.cell(row=row_start-1, column=c)
    c1.alignment = Alignment(horizontal='center') 
    c1.font = Font(bold='true')


ws_both.page_setup.orientation = ws_both.ORIENTATION_LANDSCAPE
ws_both.page_setup.paper_size = ws_both.PAPERSIZE_TABLOID
ws_both.page_setup.fitToPage = True
ws_both.page_setup.fitToHeight = False
ws_both.page_setup.fitToWidth = 1
ws_both.print_options.horizontalCentered = True
ws_both.add_print_title(6)

#Freeze Panes
c1 = ws_both.cell(row=row_start, column=2)
ws_both.freeze_panes = c1    



#Montefiore PPI's
sql = '''
        select
    sa.CustName,
    -sum(sa.TotalUnloadedCost) as Cost
from Playground.[myop\jason.walker].sales_analysis_monthly_dw sa
where
    sa.CustName in (
    'MPS-Montefiore Copier Program',
    'MPS-Montefiore Mt Vernon Copiers',
    'MPS-Montefiore New Rochelle Copiers'
    )
group by
    sa.CustName
    '''

connection = pypyodbc.connect(connection_string)
       
cur = connection.cursor()
result = cur.execute(sql).fetchall()

#Montefiore Total
sum_monte = sum(y for x, y in result)


#HiTouch
ws_cur = wb['HiTouch']
row_end = ws_ht.max_row
r = row_end + 2

x = 0
for i in range(r, r+len(result)):
    c1 = ws_cur.cell(row=i, column=1, value=result[x][0])
    c1 = ws_cur.cell(row=i, column=2, value=result[x][1])
    c1 = ws_cur.cell(row=i, column=3, value=result[x][1])
    c1 = ws_cur.cell(row=i, column=dict_bu_col.get('MPS'), value=result[x][1])
    c1 = ws_cur.cell(row=i, column=dict_bu_col.get('MPS')+1, value=result[x][1])     
    x += 1

for x in range(r, r+len(result)):
    for y in range(2, col_end+1):
        c1 = ws_cur.cell(row=x, column=y)
        c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'


#HiTouch_and_MYOP
ws_cur = wb['HiTouch_and_MYOP']
row_end = ws_ht.max_row
r = row_end + 2

x = 0
for i in range(r, r+len(result)):
    c1 = ws_cur.cell(row=i, column=1, value='[HT] '+result[x][0])
    c1 = ws_cur.cell(row=i, column=2, value=result[x][1])
    c1 = ws_cur.cell(row=i, column=3, value=result[x][1])
    c1 = ws_cur.cell(row=i, column=dict_bu_col.get('MPS'), value=result[x][1])
    c1 = ws_cur.cell(row=i, column=dict_bu_col.get('MPS')+1, value=result[x][1])     
    x += 1

for x in range(r, r+len(result)):
    for y in range(2, col_end+1):
        c1 = ws_cur.cell(row=x, column=y)
        c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'

#Other PPI's and PPC's
sql = '''
        declare @datestart date = \'{date1}\'
        declare @dateend date = \'{date2}\'
        
        select
            -sum(case when glr.level_1 = 'Sales' then gl.amount else 0 end) as Sales,
            sum(case when glr.level_1 = 'Cost' then gl.amount else 0 end) as Cost
        from TNDCSQL03.NAVRep.dbo.[Hi Touch$G_L Entry] gl with(nolock)
            inner join Playground.[myop\jason.walker].gl_account_reporting glr
                on gl.[g_l account no_] = glr.gl_account
        where
            gl.[Posting Date] between @datestart and @dateend
            and glr.company = 'HT'
            and glr.level_1 in ('Sales', 'Cost')
            and left(gl.[document no_], 3) in ('PPC', 'PPI')
        '''.format(date1=date_start, date2=date_end)
        
connection = pypyodbc.connect(connection_string)
       
cur = connection.cursor()
result = cur.execute(sql).fetchall()

#HiTouch
ws_cur = wb['HiTouch']
row_end = ws_cur.max_row
r = row_end + 1
other_ppi_ppc_sale = result[0][0] - sum_monte
c1 = ws_cur.cell(row=r, column=1, value='Other PPIs & PPCs')
c1 = ws_cur.cell(row=r, column=2, value=other_ppi_ppc_sale)
c1 = ws_cur.cell(row=r, column=3, value=result[0][1])
c1 = ws_cur.cell(row=r, column=dict_bu_col.get('OP'), value=other_ppi_ppc_sale)
c1 = ws_cur.cell(row=r, column=dict_bu_col.get('OP') + 1, value=result[0][1])

for c in range(2, col_end + 1):
    c1 = ws_cur.cell(row=r, column=c)
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'

#HiTouch and MYOP
ws_cur = wb['HiTouch_and_MYOP']
row_end = ws_cur.max_row
r = row_end + 1
c1 = ws_cur.cell(row=r, column=1, value='[HT] ' + 'Other PPIs & PPCs')
c1 = ws_cur.cell(row=r, column=2, value=other_ppi_ppc_sale)
c1 = ws_cur.cell(row=r, column=3, value=result[0][1])
c1 = ws_cur.cell(row=r, column=dict_bu_col.get('OP'), value=other_ppi_ppc_sale)
c1 = ws_cur.cell(row=r, column=dict_bu_col.get('OP') + 1, value=result[0][1])

for c in range(2, col_end + 1):
    c1 = ws_cur.cell(row=r, column=c)
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'


#Freight Revenue
sql = '''
        declare @datestart date = \'{date1}\'
        declare @dateend date = \'{date2}\'
        
        select
            cast(-sum(gl.amount) as decimal(12, 6)) as Freight_Revenue
        from TNDCSQL03.NAVRep.dbo.[MYOP$G_L Entry] gl with(nolock)
        where
            gl.[Posting Date] between @datestart and @dateend
            and gl.[g_l account no_] = '41000'
    '''.format(date1=date_start, date2=date_end)
    
connection = pypyodbc.connect(connection_string)
       
cur = connection.cursor()
result = cur.execute(sql).fetchall()
freight_rev = float(result[0][0])


#MYOP
ws_cur = wb['MYOP']
row_end = ws_cur.max_row
r = row_end + 2
c1 = ws_cur.cell(row=r, column=1, value='Freight Revenue')
c1 = ws_cur.cell(row=r, column=2, value=freight_rev)
c1 = ws_cur.cell(row=r, column=dict_bu_col.get('FUL'), value=freight_rev)
c1 = ws_cur.cell(row=r+1, column=1, value='TSC Freight Reclass')
c1 = ws_cur.cell(row=r+1, column=3, value=freight_rev*.9)
c1 = ws_cur.cell(row=r+1, column=dict_bu_col.get('FUL')+1, value=freight_rev*.9)

for x in range(r, r+2):
    for c in range(2, col_end+1):
        c1 = ws_cur.cell(row=x, column=c)
        c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'

#HiTouch_and_MYOP
ws_cur = wb['HiTouch_and_MYOP']
row_end = ws_cur.max_row
r = row_end + 1
c1 = ws_cur.cell(row=r, column=1, value='[MYOP] ' + 'Freight Revenue')
c1 = ws_cur.cell(row=r, column=2, value=freight_rev)
c1 = ws_cur.cell(row=r, column=dict_bu_col.get('FUL'), value=freight_rev)
c1 = ws_cur.cell(row=r+1, column=1, value='[MYOP] ' + 'TSC Freight Reclass')
c1 = ws_cur.cell(row=r+1, column=3, value=freight_rev*.9)
c1 = ws_cur.cell(row=r+1, column=dict_bu_col.get('FUL')+1, value=freight_rev*.9)

for x in range(r, r+2):
    for c in range(2, col_end+1):
        c1 = ws_cur.cell(row=x, column=c)
        c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'

#Adjusted Total, POS Margin %
#HiTouch
ws_cur = wb['HiTouch']

for r in range(1, ws_cur.max_row+1):
    if ws_cur.cell(row=r, column=1).value == 'DW Total':
        row_total_1 = r

row_end = ws_cur.max_row
row_next = row_end + 2      
c1 = ws_cur.cell(row=row_next, column=1, value='Adjusted Total')

for c in range(2, col_end+1):
    c1 = ws_cur.cell(row=row_next, column=c, 
                     value='=SUM({col_letter}{row1}:{col_letter}{row2})'
                     .format(col_letter=dcc.get(c), 
                             row1=row_total_1, 
                             row2=row_end))
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'

#POS Margin %
row_next += 1        
c1 = ws_cur.cell(row=row_next, column=1, value='POS Margin %')            

for c in range(3, col_end+1, 2):
    c1 = ws_cur.cell(row=row_next, column=c, 
                     value='''=IF({col_letter1}{row1}=0,0,
                     ({col_letter1}{row1}-{col_letter2}{row1})/
                     {col_letter1}{row1})'''
                     .format(col_letter1=dcc.get(c-1), 
                             row1=row_next-1, 
                             col_letter2=dcc.get(c)).replace('\n', '').replace(' ', ''))
    c1.number_format = '0.0%'
    
#MYOP
ws_cur = wb['MYOP']

for r in range(1, ws_cur.max_row+1):
    if ws_cur.cell(row=r, column=1).value == 'DW Total':
        row_total_1 = r

row_end = ws_cur.max_row
row_next = row_end + 2      
c1 = ws_cur.cell(row=row_next, column=1, value='Adjusted Total')

for c in range(2, col_end+1):
    c1 = ws_cur.cell(row=row_next, column=c, 
                     value='=SUM({col_letter}{row1}:{col_letter}{row2})'
                     .format(col_letter=dcc.get(c), 
                             row1=row_total_1, 
                             row2=row_end))
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'

#POS Margin %
row_next += 1        
c1 = ws_cur.cell(row=row_next, column=1, value='POS Margin %')            

for c in range(3, col_end+1, 2):
    c1 = ws_cur.cell(row=row_next, column=c, 
                     value='''=IF({col_letter1}{row1}=0,0,
                     ({col_letter1}{row1}-{col_letter2}{row1})/
                     {col_letter1}{row1})'''
                     .format(col_letter1=dcc.get(c-1), 
                             row1=row_next-1, 
                             col_letter2=dcc.get(c)).replace('\n', '').replace(' ', ''))
    c1.number_format = '0.0%'

#HiTouch_and_MYOP
ws_cur = wb['HiTouch_and_MYOP']

for r in range(1, ws_cur.max_row+1):
    if ws_cur.cell(row=r, column=1).value == 'DW Total':
        row_total_1 = r

row_end = ws_cur.max_row
row_next = row_end + 2      
c1 = ws_cur.cell(row=row_next, column=1, value='Adjusted Total')

for c in range(2, col_end+1):
    c1 = ws_cur.cell(row=row_next, column=c, 
                     value='=SUM({col_letter}{row1}:{col_letter}{row2})'
                     .format(col_letter=dcc.get(c), 
                             row1=row_total_1, 
                             row2=row_end))
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'

#POS Margin %
row_next += 1        
c1 = ws_cur.cell(row=row_next, column=1, value='POS Margin %')            

for c in range(3, col_end+1, 2):
    c1 = ws_cur.cell(row=row_next, column=c, 
                     value='''=IF({col_letter1}{row1}=0,0,
                     ({col_letter1}{row1}-{col_letter2}{row1})/
                     {col_letter1}{row1})'''
                     .format(col_letter1=dcc.get(c-1), 
                             row1=row_next-1, 
                             col_letter2=dcc.get(c)).replace('\n', '').replace(' ', ''))
    c1.number_format = '0.0%'    
    
#Set Page Margins
ws_cur = wb['HiTouch']
ws_cur.page_margins = PageMargins(left=.5, right=.5, top=.5, bottom=.5)    
ws_cur = wb['MYOP']
ws_cur.page_margins = PageMargins(left=.5, right=.5, top=.5, bottom=.5)
ws_cur = wb['HiTouch_and_MYOP']
ws_cur.page_margins = PageMargins(left=.5, right=.5, top=.5, bottom=.5)

    
save_path = 'c:\\temp\\'
wb.save(save_path + 'SALES ANALYSIS SUMMARY MTD {d1.year}{dmth}{dday}.xlsx'.
        format(d1=date_end,
               dmth=str(date_end.month).zfill(2),
               dday=str(date_end.day).zfill(2)))
   

print('Done')




    

